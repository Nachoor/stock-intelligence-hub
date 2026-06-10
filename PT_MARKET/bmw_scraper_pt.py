"""BMW Stock Locator Scraper — PORTUGAL
Requirements: pip install openpyxl playwright && python -m playwright install chromium
Usage: python bmw_scraper_pt.py   ->  generates STOCK_BMW_PT.xlsx
"""
from __future__ import annotations
import json, os, re, threading
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlencode
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

START_URL    = "https://www.bmw.pt/pt-pt/sl/stocklocator/results?sorting=PRICE_ASC"
OUT_XLSX     = "STOCK_BMW_PT.xlsx"
MAX_PER_PAGE = 12
MAX_VEHICLES = None

COLUMNS = [
    "Type", "Model", "Version", "Carline", "Year", "Trim", "Body",
    "Fuel", "Gearbox", "Power", "Drive", "Ext. Color", "Color Code",
    "Int. Color", "Upholstery", "RRP (EUR)", "Price Label", "Monthly Rate (EUR)",
    "APR (%)", "Nominal Rate (%)", "Deposit (EUR)", "Term (months)", "Availability",
    "Delivery Text", "Available Date", "Published Date", "Dealer", "City", "Dealer ID",
    "Dealer Email", "Car ID", "Campaign", "Online", "Business Model", "Car URL",
]

FUEL_MAP = {
    "GASOLINE":"Petrol","PETROL":"Petrol","DIESEL":"Diesel",
    "ELECTRIC":"Electric","BEV":"Electric","PHEV":"Plug-in Hybrid",
    "HEV":"Hybrid","HYBRID":"Hybrid","MHEV":"Mild Hybrid","MILD_HYBRID":"Mild Hybrid",
}
DOE_FUEL_MAP = {
    "PHEV":"Plug-in Hybrid","HEV":"Hybrid","MHEV":"Mild Hybrid",
    "BEV":"Electric","ELECTRIC":"Electric",
}
TRANS_MAP = {"AUTOMATIC":"Automatic","MANUAL":"Manual","STEPTRONIC":"Automatic Steptronic"}
DRIVE_MAP = {"FRONT":"Front-wheel","REAR":"Rear-wheel","ALL_WHEEL":"4x4 xDrive","AWD":"4x4 xDrive"}

BMW_PT_DEALER_CITY = {
    "47590":"Lisboa","32924":"Lisboa","47776":"Lisboa","50434":"Cascais","13777":"Lisboa",
    "52899":"Porto","52897":"Lisboa","52896":"Lisboa","52905":"Coimbra",
    "52875":"Braga","52894":"Funchal","52901":"Cascais","52893":"Porto",
    "02552":"Porto","34887":"Braga","52718":"Porto","52719":"Guimaraes",
    "47018":"Setubal","53289":"Montijo","36729":"Lisboa","02157":"Lisboa",
    "02181":"Almada","53067":"Setubal",
    "02575":"Braga","42447":"Evora","52255":"Faro",
    "02569":"Ponta Delgada","02546":"Funchal",
}

_PT_CITIES_RAW = [
    "Lisboa","Porto","Braga","Coimbra","Aveiro","Faro","Evora","Setubal",
    "Viseu","Leiria","Cascais","Sintra","Loures","Almada","Oeiras","Amadora",
    "Matosinhos","Gaia","Gondomar","Guimaraes","Barcelos","Maia",
    "Funchal","Portimao","Albufeira","Loule","Tomar","Torres Vedras",
    "Covilha","Fundao","Viana do Castelo","Vila Real","Braganca",
    "Seixal","Montijo","Barreiro","Espinho","Felgueiras",
    "Ponta Delgada","Evora","Faro",
]
_PT_CITIES_SORTED = sorted(set(_PT_CITIES_RAW), key=len, reverse=True)

def _pt_city_from_dealer(name):
    if not name: return ""
    for city in _PT_CITIES_SORTED:
        if re.search(r"(?i)\b" + re.escape(city) + r"\b", name):
            return city
    return ""


def text_value(v):
    if v is None: return ""
    if isinstance(v, dict):
        for k in ("pt_PT","en_GB","default","description","label","name","value"):
            if v.get(k): return text_value(v[k])
        return ""
    if isinstance(v, list): return ", ".join(text_value(i) for i in v if text_value(i))
    return re.sub(r"\s+", " ", str(v)).strip()

def deep_get(obj, *keys, default=""):
    cur = obj
    for k in keys:
        if not isinstance(cur, dict): return default
        cur = cur.get(k)
        if cur is None: return default
    return cur

def first_existing(*vals):
    for v in vals:
        if v not in (None, "", [], {}): return v
    return ""

def normalize_money(v):
    if v in (None, ""): return ""
    if isinstance(v, dict):
        v = first_existing(v.get("value"), v.get("amount"), v.get("grossAmount"))
    if isinstance(v, (int, float)): return v
    if isinstance(v, str):
        c = re.sub(r"[^\d,.-]","",v).replace(".","").replace(",",".")
        try: return float(c)
        except ValueError: return v
    return v

def build_detail_url(v, mo, fallback_car_id):
    vss_id = text_value(first_existing(v.get("vssId"), v.get("id")))
    config_id = text_value(deep_get(v, "internal", "vssConfigId"))
    if not vss_id:
        return f"https://www.bmw.pt/pt-pt/sl/stocklocator/results?vehicleId={fallback_car_id}"
    if ":" not in config_id:
        params = {"vehicleId": str(fallback_car_id)}
        return "https://www.bmw.pt/pt-pt/sl/stocklocator/details/" + vss_id + "?" + urlencode(params)

    model_code, raw_codes = config_id.split(":", 1)
    codes = [c.strip() for c in raw_codes.split(",") if c.strip()]
    paint = first_existing(*(c for c in codes if c.startswith("P")))
    fabric = first_existing(*(c for c in codes if c.startswith("F")))
    options = [c for c in codes if c not in {paint, fabric}]
    model_range_code = text_value(first_existing(
        deep_get(mo, "modelRange", "name"),
        deep_get(mo, "modelRange", "code"),
        deep_get(mo, "modelRange", "modelRangeCode"),
        v.get("modelRangeCode"),
    ))

    params = {"modelCode": model_code, "vehicleId": str(fallback_car_id)}
    if paint:
        params["paint"] = paint
    if fabric:
        params["fabric"] = fabric
    if model_range_code:
        params["modelRangeCode"] = model_range_code
    if options:
        params["options"] = ",".join(options)
    return "https://www.bmw.pt/pt-pt/sl/stocklocator/details/" + vss_id + "?" + urlencode(params, safe=",")

def accept_cookies(page):
    for label in ["Accept all","Aceitar tudo","Aceitar","Allow all","Confirmar"]:
        try:
            loc = page.get_by_text(label, exact=False).first
            if loc.count() > 0:
                loc.click(timeout=3000)
                page.wait_for_timeout(1000)
                print("   Cookies accepted")
                return
        except Exception:
            continue
    for selector in ["button#onetrust-accept-btn-handler","button.accept-all",
                     "[data-testid='accept-all-cookies']","button[class*='accept']"]:
        try:
            el = page.locator(selector).first
            if el.count() > 0:
                el.click(timeout=3000)
                page.wait_for_timeout(1000)
                print(f"   Cookies accepted ({selector})")
                return
        except Exception:
            continue
    print("   Warning: cookie banner not found")

def map_vehicle(hit, pricing_map=None):
    v = hit.get("vehicle") or hit.get("document") or hit.get("data") or hit
    if not isinstance(v, dict): return {}
    vspec   = v.get("vehicleSpecification") or {}
    mo      = vspec.get("modelAndOption") or v.get("modelAndOption") or {}
    tech    = ((vspec.get("technicalAndEmission") or {}).get("technicalData") or {})
    order   = v.get("ordering") or {}
    retail  = order.get("retailData") or {}
    distrib = order.get("distributionData") or {}
    # pricing: en v, vspec, hit, o pricing_map por vehicle ID
    pricing = v.get("pricing") or vspec.get("pricing") or hit.get("pricing") or {}
    car_id_for_lookup = str(first_existing(
        v.get("documentId"), v.get("id"), hit.get("id"), hit.get("documentId"), ""))
    if not pricing and pricing_map and car_id_for_lookup:
        pricing = pricing_map.get(car_id_for_lookup, {})
    price   = pricing.get("price") or pricing.get("retailPrice") or {}
    install = pricing.get("monthlyInstallment") or {}
    calcs   = deep_get(install, "selectedSfOffer", "calculations", default=[])
    calc    = calcs[0] if isinstance(calcs, list) and calcs else {}

    model_range = text_value(deep_get(mo, "modelRange", "description"))
    minfo       = mo.get("model") or {}
    derivative  = text_value(first_existing(
        minfo.get("derivative"), minfo.get("modelName"), mo.get("modelName")))
    modelo      = " ".join(p for p in [model_range, derivative] if p).strip()

    power    = tech.get("power") or {}
    kw       = first_existing(power.get("maxPowerKW"), power.get("kw"))
    ps_      = first_existing(power.get("maxPowerPS"), power.get("ps"), power.get("hp"))
    potencia = f"{kw} kW ({ps_} HP)" if kw and ps_ else (f"{kw} kW" if kw else "")

    exp_date  = first_existing(
        distrib.get("expectedDeliveryDate"),
        deep_get(distrib,"earliestRealHandover","date"))
    today_str = date.today().isoformat()
    if exp_date:
        try:
            disponible = (
                "available"
                if datetime.fromisoformat(str(exp_date)[:10]).date() <= date.today()
                else "soon"
            )
        except ValueError:
            disponible = "soon"
    else:
        disponible = ""

    sales_dest = deep_get(v, "offering", "salesDestinations", default=[])

    city = first_existing(
        deep_get(distrib, "locationOutletAddress", "city"),
        deep_get(retail,  "locationOutletAddress", "city"),
        deep_get(distrib, "locationOutletAddress", "municipality"),
        deep_get(retail,  "locationOutletAddress", "municipality"),
        deep_get(v, "dealer", "city"),
    )
    dealer_id_raw = text_value(first_existing(
        retail.get("buNo"), distrib.get("buNo"), v.get("buNo")))
    if not city and dealer_id_raw:
        city = BMW_PT_DEALER_CITY.get(str(dealer_id_raw), "")
    if not city:
        dealer_name_raw = text_value(first_existing(
            retail.get("locationOutletName"),
            distrib.get("destinationLocationDomesticDealerName"),
        ))
        city = _pt_city_from_dealer(dealer_name_raw)

    car_id = first_existing(
        v.get("documentId"), v.get("id"), hit.get("id"), hit.get("documentId"))
    published_date = first_existing(
        v.get("datePublished"),
        v.get("publicationDate"),
        v.get("publishedDate"),
        v.get("listingDate"),
        v.get("onlineSince"),
        deep_get(v, "offering", "datePublished"),
        deep_get(v, "offering", "publicationDate"),
        deep_get(v, "offering", "publishedDate"),
        deep_get(v, "offering", "listingDate"),
        deep_get(v, "offering", "onlineSince"),
    )

    rrp = normalize_money(first_existing(
        price.get("value"), price.get("amount"), price.get("grossAmount"),
        price.get("grossValue"), price.get("netValue"),
        deep_get(pricing, "retailPrice", "value"),
        deep_get(pricing, "retailPrice", "amount"),
        deep_get(pricing, "retailPrice", "grossAmount"),
        deep_get(pricing, "retailPrice", "grossValue"),
        deep_get(pricing, "listPrice", "value"),
        deep_get(pricing, "listPrice", "grossAmount"),
        deep_get(pricing, "listPrice", "grossValue"),
        deep_get(pricing, "configuredPrice", "value"),
        deep_get(pricing, "configuredPrice", "grossAmount"),
        deep_get(pricing, "totalPrice", "value"),
        deep_get(pricing, "totalPrice", "grossAmount"),
        deep_get(pricing, "pvp", "value"), deep_get(pricing, "pvp", "amount"),
        deep_get(v, "price", "value"), deep_get(v, "price", "amount"),
        deep_get(v, "price", "grossAmount"), deep_get(v, "retailPrice"),
        deep_get(v, "listPrice", "value"),
        deep_get(hit, "price", "value"), deep_get(hit, "price", "amount"),
        deep_get(hit, "retailPrice"),
    ))

    return {
        "Type":       "New",
        "Model":      modelo,
        "Version":    derivative,
        "Carline":    model_range,
        "Year":       text_value(deep_get(minfo,"effectDateRange","from"))[:4],
        "Trim":       text_value(mo.get("line")),
        "Body":       text_value(first_existing(
                          deep_get(mo,"bodyTypeDescription","pt_PT"),
                          mo.get("bodyType"))),
        "Fuel":       (lambda doe,bf: DOE_FUEL_MAP.get(doe) or FUEL_MAP.get(bf) or bf)(
                          text_value(tech.get("degreeOfElectrificationBasedFuelType") or ""),
                          text_value(mo.get("baseFuelType") or ""),
                      ),
        "Gearbox":    TRANS_MAP.get(
                          text_value(mo.get("transmission")),
                          text_value(mo.get("transmission"))),
        "Power":      potencia,
        "Drive":      DRIVE_MAP.get(
                          text_value(mo.get("driveType")),
                          text_value(mo.get("driveType"))),
        "Ext. Color": text_value(first_existing(
                          deep_get(mo,"color","clusterFine"),
                          deep_get(mo,"color","description"))),
        "Color Code": text_value(first_existing(
                          mo.get("paintCode"),
                          deep_get(mo,"color","hexColorCode"))),
        "Int. Color": text_value(deep_get(mo,"upholsteryColor","upholsteryColorCluster")),
        "Upholstery": text_value(mo.get("upholsteryType")),
        "RRP (EUR)":  rrp,
        "Price Label": text_value(
                          price.get("label") or
                          deep_get(pricing,"retailPrice","label")),
        "Monthly Rate (EUR)": normalize_money(calc.get("monthlyRate")),
        "APR (%)":    calc.get("effectiveAnnualInterestRate") or "",
        "Nominal Rate (%)": calc.get("nominalInterestRate") or "",
        "Deposit (EUR)": normalize_money(calc.get("deposit")),
        "Term (months)": calc.get("term") or "",
        "Availability":  disponible,
        "Delivery Text": f"Expected delivery: {exp_date}" if exp_date else "",
        "Available Date": exp_date or today_str,
        "Published Date": published_date,
        "Dealer":     text_value(first_existing(
                          retail.get("locationOutletName"),
                          distrib.get("destinationLocationDomesticDealerName"))),
        "City":       text_value(city),
        "Dealer ID":  dealer_id_raw,
        "Dealer Email": text_value(distrib.get("dealerEmail")),
        "Car ID":     text_value(car_id),
        "Campaign":   "",
        "Online":     "true" if "ONLINE" in sales_dest else "false",
        "Business Model": text_value(
                          deep_get(v,"salesProcess","type",default="dealer_stock")),
        "Car URL":    build_detail_url(v, mo, car_id),
    }

def save_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cars in stock"
    hfill = PatternFill("solid", fgColor="1C69D4")
    hfont = Font(bold=True, color="FFFFFF")
    for ci, cn in enumerate(COLUMNS, 1):
        cell = ws.cell(1, ci, cn)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, cn in enumerate(COLUMNS, 1):
            ws.cell(ri, ci, row.get(cn, ""))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    money = {"RRP (EUR)","Monthly Rate (EUR)","Deposit (EUR)"}
    for wr in ws.iter_rows(min_row=2):
        for cell in wr:
            if COLUMNS[cell.column-1] in money and isinstance(cell.value,(int,float)):
                cell.number_format = '#,##0.00 "EUR"'
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml+2, 55)
    wb.save(output_path)


def _try_click_load_more(page):
    """Click the BMW PT 'Mostrar mais' NEO-BUTTON web component."""
    # BMW PT uses a <neo-button> web component — JS click is most reliable
    try:
        clicked = page.evaluate("""
            (() => {
                // Primary: NEO-BUTTON with btn-goto-results class (BMW PT stock locator)
                const nb = document.querySelector('neo-button.btn-goto-results');
                if (nb) { nb.click(); return 'neo-button.btn-goto-results'; }
                // Fallback: any neo-button with role=button near pagination
                const pag = document.querySelector('.results-pagination-endless');
                if (pag) {
                    const neo = pag.querySelector('neo-button');
                    if (neo) { neo.click(); return 'pag neo-button'; }
                }
                // Fallback: text match in shadow DOM
                for (const el of document.querySelectorAll('neo-button,[role="button"]')) {
                    const t = el.innerText || el.textContent || '';
                    if (t.includes('Mostrar') || t.includes('mais') || t.includes('Load more')) {
                        el.click(); return 'text-match: ' + t.slice(0,20);
                    }
                }
                return null;
            })()
        """)
        if clicked:
            return True
    except Exception:
        pass

    # Playwright locator fallbacks
    for sel in [
        "neo-button.btn-goto-results",
        ".results-pagination-endless neo-button",
        "[role='button']:near(.results-pagination-endless)",
        "neo-button[class*='btn-goto']",
    ]:
        try:
            btn = page.locator(sel).first
            if btn.count() > 0:
                btn.scroll_into_view_if_needed()
                page.wait_for_timeout(200)
                btn.click(timeout=3000)
                return True
        except Exception:
            pass
    return False


def scrape_bmw():
    from playwright.sync_api import sync_playwright
    all_hits    = []
    total_count = [0]
    price_count = [0]   # hits con precio detectado
    seen_ids    = set()
    lock        = threading.Lock()
    # Map vehicle_id -> pricing dict (captured from any pricing-related endpoint)
    pricing_map = {}

    def on_response(response):
        url = response.url
        try:
            ct = response.headers.get("content-type","")
            if "json" not in ct:
                return
        except Exception:
            return

        # ---- 1. Vehiclesearch/search: captura hits (incluye pricing cuando el browser renderiza)
        if "vehiclesearch" in url and "/search/" in url:
            try:
                data = response.json()
                t    = (data.get("metadata") or {}).get("totalCount", 0)
                hits = data.get("hits") or []
                with lock:
                    if t and t > total_count[0]:  # nunca sobreescribir con valor menor
                        total_count[0] = t
                    for h in hits:
                        hid = h.get("documentId") or h.get("id") or id(h)
                        if hid not in seen_ids:
                            seen_ids.add(hid)
                            all_hits.append(h)
                            # Detectar precio rápidamente sin parsear completo
                            _v = h.get("vehicle") or h.get("document") or h.get("data") or h
                            _pr = (
                                _v.get("pricing") or
                                (_v.get("vehicleSpecification") or {}).get("pricing") or
                                h.get("pricing") or {}
                            )
                            _p = _pr.get("price") or _pr.get("retailPrice") or _pr.get("listPrice") or {}
                            if _p.get("value") or _p.get("amount") or _p.get("grossAmount"):
                                price_count[0] += 1
                if hits:
                    total = total_count[0] or 1
                    pct   = round(len(all_hits)*100/total)
                    ppct  = round(price_count[0]*100/len(all_hits)) if all_hits else 0
                    print(
                        f"\r   Captured: {len(all_hits)}/{total_count[0]} ({pct}%)  "
                        f"| Prices: {price_count[0]} ({ppct}%)",
                        end="", flush=True)
            except Exception:
                pass
            return

        # ---- 2. Cualquier endpoint de pricing independiente
        price_keywords = ("pric","financ","quot","install","rate","offer","leasing")
        if not any(kw in url.lower() for kw in price_keywords):
            return
        try:
            data = response.json()
            if not isinstance(data, dict):
                return
            # Extraer vehicle ID de la URL o del body
            vid = None
            for pattern in [
                r'/([A-Z0-9]{12,})(?:[/?]|$)',
                r'vehicleId=([A-Z0-9]{10,})',
                r'documentId=([A-Z0-9]{10,})',
            ]:
                m = re.search(pattern, url)
                if m:
                    vid = m.group(1)
                    break
            if not vid:
                vid = str(first_existing(
                    data.get("vehicleId"), data.get("documentId"),
                    data.get("id"),
                    deep_get(data,"vehicle","id"),
                    deep_get(data,"vehicle","documentId"),
                ) or "")
            if vid:
                with lock:
                    pricing_map[vid] = data
        except Exception:
            pass

    # Capture API request for POST fallback
    captured_req = [None]

    def on_request(request):
        url = request.url
        is_search    = "vehiclesearch" in url and "/search/"    in url
        is_aggregate = "vehiclesearch" in url and "/aggregate/" in url
        if not (is_search or is_aggregate):
            return
        try:
            body = request.post_data_json
            if body is None:
                raw = request.post_data
                if raw:
                    body = json.loads(raw)
            if not body:
                return
            base_url = url.split("?")[0]
            if is_search or captured_req[0] is None:
                if is_aggregate:
                    base_url = base_url.replace("/aggregate/", "/search/")
                captured_req[0] = {"url": base_url, "body": body}
        except Exception:
            pass

    with sync_playwright() as p:
        print("Opening browser...")
        headless = os.getenv("BMW_HEADLESS", "0").lower() in ("1", "true", "yes")
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="pt-PT",
            viewport={"width": 1440, "height": 900},
        )
        page = ctx.new_page()
        page.on("request", on_request)
        page.on("response", on_response)
        page.set_default_timeout(120000)

        print("Loading BMW Portugal Stock Locator...")
        # First load to get the domain, then clear stored filters
        page.goto("https://www.bmw.pt", wait_until="domcontentloaded", timeout=30000)
        page.evaluate(
            "try { localStorage.clear(); sessionStorage.clear(); } catch(e) {}"
        )
        # Clear cookies to remove any server-side session filters
        ctx.clear_cookies()
        # Now load the stock locator clean (no stored model filters)
        page.goto(START_URL, wait_until="domcontentloaded", timeout=120000)
        accept_cookies(page)
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass

        print("Waiting for first API response...")
        for _ in range(80):
            if total_count[0] > 0:
                break
            page.wait_for_timeout(500)

        if total_count[0] == 0:
            browser.close()
            raise RuntimeError("No API search calls detected.")

        total = total_count[0]
        print("   Total vehicles: " + str(total))
        page.wait_for_timeout(3000)

        # ---- browser fetch() para TODOS los vehiculos (con precio) ----
        # Usamos el fetch() nativo del browser: incluye sesion/cookies → precios.
        # Lanzamos en paralelo grupos de BATCH_SIZE requests y esperamos las respuestas.
        if captured_req[0]:
            import copy
            remaining_starts = list(range(
                len(all_hits) if len(all_hits) % MAX_PER_PAGE == 0
                else (len(all_hits) // MAX_PER_PAGE + 1) * MAX_PER_PAGE,
                total, MAX_PER_PAGE,
            ))
            req      = captured_req[0]
            base_url = req["url"]
            body     = req["body"]
            BATCH_SIZE = 8   # requests en paralelo por grupo
            BATCH_WAIT = 2500  # ms para esperar respuestas del grupo
            POLL_MS    = 150

            print("   Fetching " + str(len(remaining_starts)) + " pages via browser fetch()...")

            for i in range(0, len(remaining_starts), BATCH_SIZE):
                if MAX_VEHICLES and len(all_hits) >= MAX_VEHICLES:
                    break
                batch = remaining_starts[i:i+BATCH_SIZE]
                prev  = len(all_hits)

                # Lanzar todas las requests del grupo en paralelo via browser fetch()
                for start in batch:
                    paged = copy.deepcopy(body)
                    paged["startIndex"] = start
                    paged["maxResults"] = MAX_PER_PAGE
                    for key in ("query","filter","pagination","searchRequest","request"):
                        if key in paged and isinstance(paged[key], dict):
                            paged[key]["startIndex"] = start
                            paged[key]["maxResults"] = MAX_PER_PAGE
                    url_p = (base_url + "?maxResults=" + str(MAX_PER_PAGE)
                             + "&startIndex=" + str(start) + "&brand=BMW")
                    js_body = json.dumps(paged)
                    # fire-and-forget: on_response captura la respuesta
                    try:
                        page.evaluate(
                            "(a) => { fetch(a[0],{method:'POST',"
                            "headers:{'Content-Type':'application/json'},"
                            "body:a[1]}).catch(function(){}); }",
                            [url_p, js_body]
                        )
                    except Exception:
                        pass

                # Esperar que lleguen las respuestas del grupo
                deadline = BATCH_WAIT // POLL_MS
                for _ in range(deadline):
                    page.wait_for_timeout(POLL_MS)
                    if len(all_hits) >= prev + len(batch) * MAX_PER_PAGE * 0.75:
                        break

                pct  = round(len(all_hits) * 100 / total)
                ppct = round(price_count[0] * 100 / len(all_hits)) if all_hits else 0
                print("   Captured: " + str(len(all_hits)) + "/" + str(total)
                      + " (" + str(pct) + "%)  | Prices: "
                      + str(price_count[0]) + " (" + str(ppct) + "%)")

        else:
            print("   Warning: no captured request — only initial hits available.")
        print("   Parsing " + str(len(all_hits)) + " vehicles...")
        if pricing_map:
            print("   Pricing map entries: " + str(len(pricing_map)))
        browser.close()

    rows = [map_vehicle(h, pricing_map) for h in all_hits]
    return [r for r in rows if r]



def main():
    sep = "=" * 65
    start = datetime.now()
    print(sep)
    print("  BMW PORTUGAL -- STOCK SCRAPER")
    print("  " + start.strftime("%Y-%m-%d %H:%M:%S"))
    print("  Source: bmw.pt Stock Locator (vehiclesearch API)")
    print(sep)
    rows = scrape_bmw()
    elapsed = datetime.now() - start
    save_excel(rows, OUT_XLSX)
    n = len(rows)
    xl = str(Path(OUT_XLSX).resolve())
    wp = sum(1 for r in rows if r.get("RRP (EUR)"))
    wc = sum(1 for r in rows if r.get("City"))
    pp = round(wp * 100 / max(n, 1))
    pc = round(wc * 100 / max(n, 1))
    print("")
    print(sep)
    print("  SUMMARY")
    print("  Vehicles  : " + str(n))
    print("  With price: " + str(wp) + " (" + str(pp) + "%)")
    print("  With city : " + str(wc) + " (" + str(pc) + "%)")
    print("  Excel     : " + xl)
    print("  Time      : " + str(elapsed))
    print(sep)


if __name__ == "__main__":
    main()
