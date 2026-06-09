"""BMW Stock Locator Scraper - España
Requisitos: pip install openpyxl playwright && python -m playwright install chromium
Uso: python bmw_scraper.py  ->  genera STOCK_BMW.xlsx
"""
from __future__ import annotations
import json, os, random, re, threading
from datetime import date, datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# URL sin filtros: recoge TODOS los vehículos del stock locator
START_URL    = "https://www.bmw.es/es-es/sl/stocklocator/results?sorting=PRICE_ASC"
OUT_XLSX     = "STOCK_BMW.xlsx"
PARTIAL_XLSX = "STOCK_BMW.partial.xlsx"
MAX_PER_PAGE = 12   # Máximo que acepta la API (>12 devuelve 400). Hash no es necesario.
MAX_VEHICLES = int(os.getenv("BMW_MAX_VEHICLES", "0") or "0") or None

COLUMNS = [
    "Tipo","Modelo","Version","Carline","Ano","Trim","Carroceria",
    "Combustible","Cambio","Potencia","Traccion","Color Ext.","Codigo Color",
    "Color Int.","Tapizado","PVP (EUR)","Etiqueta Precio","Cuota/mes (EUR)",
    "TAE (%)","TIN (%)","Entrada (EUR)","Plazo (meses)","Disponibilidad",
    "Texto Entrega","Fecha Disponible","Fecha Publicacion","Concesionario","Ciudad","ID Dealer",
    "Email Dealer","Car ID","Campana","Online","Modelo Negocio","URL Coche",
]
FUEL_MAP = {"GASOLINE":"Gasolina","PETROL":"Gasolina","DIESEL":"Diesel",
            "ELECTRIC":"Electrico","BEV":"Electrico","PHEV":"Hibrido enchufable",
            "HEV":"Hibrido","HYBRID":"Hibrido","MHEV":"Hibrido suave","MILD_HYBRID":"Hibrido suave"}
DOE_FUEL_MAP = {"PHEV":"Hibrido enchufable","HEV":"Hibrido","MHEV":"Hibrido suave",
                "BEV":"Electrico","ELECTRIC":"Electrico"}
TRANS_MAP = {"AUTOMATIC":"Automatico","MANUAL":"Manual","STEPTRONIC":"Automatico Steptronic"}
DRIVE_MAP = {"FRONT":"Delantera","REAR":"Trasera","ALL_WHEEL":"4x4 xDrive","AWD":"4x4 xDrive"}

def text_value(v):
    if v is None: return ""
    if isinstance(v, dict):
        for k in ("es_ES","default_ES","es","default","description","label","name","value"):
            if v.get(k): return text_value(v[k])
        return ""
    if isinstance(v, list): return ", ".join(text_value(i) for i in v if text_value(i))
    return re.sub(r"\s+", " ", str(v)).strip()

def deep_get(obj, *keys, default=""):
    cur = obj
    for k in keys:
        if not isinstance(cur, dict): return default
        cur = cur.get(k)
        if cur is None: return default
    return cur

def first_existing(*vals):
    for v in vals:
        if v not in (None, "", [], {}): return v
    return ""

def hit_id(hit):
    if not isinstance(hit, dict):
        return str(id(hit))
    v = hit.get("vehicle") or hit.get("document") or hit.get("data") or {}
    if not isinstance(v, dict):
        v = {}
    return str(first_existing(
        hit.get("documentId"),
        hit.get("id"),
        v.get("documentId"),
        v.get("id"),
        id(hit),
    ))

def normalize_money(v):
    if v in (None, ""): return ""
    if isinstance(v, dict): v = first_existing(v.get("value"), v.get("amount"), v.get("grossAmount"))
    if isinstance(v, (int, float)): return v
    if isinstance(v, str):
        c = re.sub(r"[^\d,.-]","",v).replace(".","").replace(",",".")
        try: return float(c)
        except ValueError: return v
    return v

def accept_cookies(page):
    for label in ["Aceptar todo","Aceptar todas","Aceptar cookies","Aceptar","Confirmar",
                  "Accept all","Accept cookies","Allow all","Alle akzeptieren"]:
        try:
            loc = page.get_by_text(label, exact=False).first
            if loc.count() > 0:
                loc.click(timeout=3000)
                page.wait_for_timeout(1000)
                print("   Cookies aceptadas")
                return
        except Exception:
            continue
    for selector in ["button#onetrust-accept-btn-handler","button.accept-all",
                     "[data-testid='accept-all-cookies']","button[class*='accept']",
                     "button[id*='accept']",".cookie-accept",".cookies-accept"]:
        try:
            el = page.locator(selector).first
            if el.count() > 0:
                el.click(timeout=3000)
                page.wait_for_timeout(1000)
                print("   Cookies aceptadas (CSS)")
                return
        except Exception:
            continue
    print("   Aviso: no se encontro banner de cookies")

_BMW_ES_DEALER_CITIES = {
    "Momentum Norte":"San Sebastian","Ceres Motor":"Caceres","Lugauto":"Lugo",
    "Auto Premier Baviera":"Leganes","Adler Motor":"Toledo","Motri Motor":"Motril",
    "Carteya Motor":"Algeciras","San Rafael Motor":"Cordoba","Ilbira Motor":"Granada",
    "Avilcar":"Avila","Movitransa Cars":"Jerez de la Frontera","San Pablo Motor":"Sevilla",
    "Albamocion":"Albacete","Augusta Aragon":"Zaragoza","Autoram":"Zamora",
    "Grunblau Motor":"Santander","Lizaga Motor":"Pamplona","Vehinter":"Getafe",
    "Celtamotor":"Vigo","Movilnorte":"Madrid","Enekuri Motor":"Bilbao",
    "Engasa":"Valencia","Bernesga Motor":"Leon","Fuenteolid":"Valladolid",
    "Amiocar":"A Coruna","Automotor Costa":"Almeria","Automotor Premium":"Marbella",
    "Hispamovil":"Torrevieja","Autoberon":"Logrono","Altrac Europa":"Barcelona",
    "Caetano Cuzco":"Madrid","Proa Premium":"Ibiza","Movil Begar Levante":"Alicante",
    "Burgocar":"Burgos","Tormes Motor":"Salamanca","Autogal":"A Coruna",
    "Cabrero Motorsport":"Huesca","Oliva Motor Gavarres":"Tarragona",
    "Triocar Comercial":"Almeria","Automoviles Bertolin":"Valencia",
    "Pruna Motor":"Barcelona","Automotor":"Malaga","Automoviles Fersan":"Alicante",
}

_ES_CITIES = sorted([
    "Madrid","Barcelona","Valencia","Sevilla","Zaragoza","Malaga","Murcia",
    "Palma","Las Palmas","Bilbao","Alicante","Cordoba","Valladolid","Vigo",
    "Gijon","A Coruna","Granada","Vitoria","Elche","Oviedo","Badalona",
    "Cartagena","Terrassa","Jerez","Pamplona","Almeria","Leganes","Santander",
    "Burgos","Albacete","Getafe","Salamanca","Logrono","Huelva","Badajoz",
    "San Sebastian","Tarragona","Leon","Cadiz","Marbella","Algeciras",
    "Torrevieja","Girona","Lugo","Motril","Valladolid","Zamora","Huesca",
    "Ibiza","Toledo","Avila","Caceres","Talavera de la Reina",
], key=len, reverse=True)

def _city_from_dealer(name):
    if not name: return ""
    n = re.sub(r"\s+", " ", name).strip()
    import unicodedata
    def strip_acc(s):
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    nl = strip_acc(n.lower())
    for key, city in _BMW_ES_DEALER_CITIES.items():
        if strip_acc(key.lower()) in nl:
            return city
    for city in _ES_CITIES:
        if re.search(r"(?i)\b" + re.escape(strip_acc(city)) + r"\b", nl):
            return city
    return ""

def map_vehicle(hit, pricing_map=None):
    v = hit.get("vehicle") or hit.get("document") or hit.get("data") or hit
    if not isinstance(v, dict): return {}
    vspec   = v.get("vehicleSpecification") or {}
    mo      = vspec.get("modelAndOption") or v.get("modelAndOption") or {}
    tech    = ((vspec.get("technicalAndEmission") or {}).get("technicalData") or {})
    order   = v.get("ordering") or {}
    retail  = order.get("retailData") or {}
    distrib = order.get("distributionData") or {}
    car_id  = str(first_existing(v.get("documentId"), v.get("id"), hit.get("id"), hit.get("documentId")) or "")
    # Pricing: v.pricing.price.value es el campo real de BMW ES
    pricing = v.get("pricing") or vspec.get("pricing") or hit.get("pricing") or {}
    if not pricing and pricing_map and car_id:
        pricing = pricing_map.get(car_id, {})
    price   = pricing.get("price") or pricing.get("retailPrice") or {}
    install = pricing.get("monthlyInstallment") or {}
    calcs   = deep_get(install, "selectedSfOffer", "calculations", default=[])
    calc    = calcs[0] if isinstance(calcs, list) and calcs else {}
    model_range = text_value(deep_get(mo, "modelRange", "description"))
    minfo       = mo.get("model") or {}
    derivative  = text_value(first_existing(minfo.get("derivative"), minfo.get("modelName"), mo.get("modelName")))
    modelo      = " ".join(p for p in [model_range, derivative] if p).strip()
    power       = tech.get("power") or {}
    kw          = first_existing(power.get("maxPowerKW"), power.get("kw"))
    ps_         = first_existing(power.get("maxPowerPS"), power.get("ps"), power.get("hp"))
    potencia    = (str(kw) + " kW (" + str(ps_) + " CV)") if kw and ps_ else (str(kw) + " kW" if kw else "")
    exp_date    = first_existing(distrib.get("expectedDeliveryDate"), deep_get(distrib,"earliestRealHandover","date"))
    today_str   = date.today().isoformat()
    if exp_date:
        try: disponible = "available" if datetime.fromisoformat(str(exp_date)[:10]).date() <= date.today() else "soon"
        except ValueError: disponible = "soon"
    else: disponible = ""
    sales_dest = deep_get(v, "offering", "salesDestinations", default=[])
    city = first_existing(
        deep_get(distrib, "locationOutletAddress", "city"),
        deep_get(retail,  "locationOutletAddress", "city"),
        deep_get(distrib, "locationOutletAddress", "municipality"),
        deep_get(retail,  "locationOutletAddress", "municipality"),
        deep_get(v, "dealer", "city"),
    )
    if not city:
        dealer_name = text_value(first_existing(
            retail.get("locationOutletName"),
            distrib.get("destinationLocationDomesticDealerName"),
        ))
        city = _city_from_dealer(dealer_name)
    dealer_id = text_value(first_existing(retail.get("buNo"), distrib.get("buNo"), v.get("buNo")))
    offer_prices = deep_get(v, "offering", "offerPrices", default={})
    offer_price = {}
    if isinstance(offer_prices, dict) and offer_prices:
        offer_price = offer_prices.get(str(dealer_id)) if dealer_id else None
        if not isinstance(offer_price, dict):
            offer_price = next((op for op in offer_prices.values() if isinstance(op, dict)), {})
    fecha_publicacion = first_existing(
        v.get("publishedDate"),
        v.get("publicationDate"),
        v.get("listingDate"),
        v.get("onlineSince"),
        deep_get(v, "offering", "publishedDate"),
        deep_get(v, "offering", "publicationDate"),
        deep_get(v, "offering", "listingDate"),
        deep_get(v, "offering", "onlineSince"),
        offer_price.get("offerPriceCreatedAt") if isinstance(offer_price, dict) else "",
    )
    # PVP: v.pricing.price.value es el campo verificado para BMW ES
    pvp = normalize_money(first_existing(
        price.get("value"), price.get("amount"), price.get("grossAmount"),
        price.get("grossValue"), price.get("netValue"),
        deep_get(pricing, "retailPrice", "value"),
        deep_get(pricing, "retailPrice", "grossAmount"),
        deep_get(pricing, "listPrice", "value"),
        deep_get(pricing, "configuredPrice", "value"),
        deep_get(pricing, "totalPrice", "value"),
        deep_get(v, "price", "grossSalesPrice"),
        deep_get(v, "price", "grossListPrice"),
        deep_get(hit, "price", "value"), deep_get(hit, "retailPrice"),
    ))
    return {
        "Tipo":"Nuevo","Modelo":modelo,"Version":derivative,"Carline":model_range,
        "Ano":text_value(deep_get(minfo,"effectDateRange","from"))[:4],
        "Trim":text_value(mo.get("line")),
        "Carroceria":text_value(first_existing(deep_get(mo,"bodyTypeDescription","es_ES"), mo.get("bodyType"))),
        "Combustible":(lambda doe, bf: DOE_FUEL_MAP.get(doe) or FUEL_MAP.get(bf) or bf)(
            text_value(tech.get("degreeOfElectrificationBasedFuelType") or ""),
            text_value(mo.get("baseFuelType") or ""),
        ),
        "Cambio":TRANS_MAP.get(text_value(mo.get("transmission")), text_value(mo.get("transmission"))),
        "Potencia":potencia,
        "Traccion":DRIVE_MAP.get(text_value(mo.get("driveType")), text_value(mo.get("driveType"))),
        "Color Ext.":text_value(first_existing(deep_get(mo,"color","clusterFine"), deep_get(mo,"color","description"))),
        "Codigo Color":text_value(first_existing(mo.get("paintCode"), deep_get(mo,"color","hexColorCode"))),
        "Color Int.":text_value(deep_get(mo,"upholsteryColor","upholsteryColorCluster")),
        "Tapizado":text_value(mo.get("upholsteryType")),
        "PVP (EUR)":pvp,
        "Etiqueta Precio":text_value(price.get("label") or deep_get(pricing,"retailPrice","label")),
        "Cuota/mes (EUR)":normalize_money(calc.get("monthlyRate")),
        "TAE (%)":calc.get("effectiveAnnualInterestRate") or "",
        "TIN (%)":calc.get("nominalInterestRate") or "",
        "Entrada (EUR)":normalize_money(calc.get("deposit")),
        "Plazo (meses)":calc.get("term") or "",
        "Disponibilidad":disponible,
        "Texto Entrega":"Entrega aproximada: " + str(exp_date) if exp_date else "",
        "Fecha Disponible":exp_date or today_str,
        "Fecha Publicacion":fecha_publicacion,
        "Concesionario":text_value(first_existing(retail.get("locationOutletName"), distrib.get("destinationLocationDomesticDealerName"))),
        "Ciudad":text_value(city),
        "ID Dealer":dealer_id,
        "Email Dealer":text_value(distrib.get("dealerEmail")),
        "Car ID":text_value(car_id),
        "Campana":"",
        "Online":"true" if "ONLINE" in sales_dest else "false",
        "Modelo Negocio":text_value(deep_get(v,"salesProcess","type",default="dealer_stock")),
        "URL Coche":"https://www.bmw.es/es-es/sl/stocklocator/results?vehicleId=" + str(car_id),
    }

def save_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coches en stock"
    hfill = PatternFill("solid", fgColor="1C69D4")
    hfont = Font(bold=True, color="FFFFFF")
    for ci, cn in enumerate(COLUMNS, 1):
        cell = ws.cell(1, ci, cn)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for ri, row in enumerate(rows, 2):
        for ci, cn in enumerate(COLUMNS, 1):
            ws.cell(ri, ci, row.get(cn, ""))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    money = {"PVP (EUR)","Cuota/mes (EUR)","Entrada (EUR)"}
    for wr in ws.iter_rows(min_row=2):
        for cell in wr:
            if COLUMNS[cell.column-1] in money and isinstance(cell.value,(int,float)):
                cell.number_format = '#,##0.00 "EUR"'
    for col in ws.columns:
        ml = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml+2, 55)
    wb.save(output_path)

def scrape_bmw():
    from playwright.sync_api import sync_playwright
    import urllib.parse as _up
    import urllib.request as _urlreq
    import urllib.error as _urlerr
    import concurrent.futures as _cf

    all_hits   = []
    total_count = [0]
    price_count = [0]
    captured_req = [None]
    aggregate_model_ranges = []
    seen_ids = set()
    lock = threading.Lock()
    pricing_map = {}
    _SKIP_HDR = {"host","content-length","connection","accept-encoding"}

    def extract_marketing_model_ranges(aggregate_payload):
        found = {}

        def walk(node):
            if isinstance(node, dict):
                if node.get("name") == "MARKETING_MODEL_RANGE":
                    for value in node.get("values") or []:
                        if isinstance(value, dict) and value.get("name"):
                            found[str(value["name"])] = int(value.get("count") or 0)
                for child in node.get("values") or []:
                    walk(child)
            elif isinstance(node, list):
                for child in node:
                    walk(child)

        walk((aggregate_payload or {}).get("aggregateGeneric") or [])
        return found

    def on_request(request):
        url = request.url
        is_search    = "vehiclesearch" in url and "/search/"    in url
        is_aggregate = "vehiclesearch" in url and "/aggregate/" in url
        if not (is_search or is_aggregate): return
        try:
            body = request.post_data_json
            if body is None:
                raw = request.post_data
                if raw: body = json.loads(raw)
            if not body: return
            base_url = url.split("?")[0]
            raw_params = dict(_up.parse_qsl(url.split("?",1)[1] if "?" in url else ""))
            # Solo mantener brand y context — hash puede expirar durante la sesion
            fixed_params = {k: v for k, v in raw_params.items()
                            if k.lower() in ("brand","context")}
            hdrs = {k: v for k, v in (request.headers or {}).items()
                    if k.lower() not in _SKIP_HDR}
            if is_search or captured_req[0] is None:
                if is_aggregate:
                    base_url = base_url.replace("/aggregate/", "/search/")
                captured_req[0] = {
                    "url": base_url,
                    "headers": hdrs,
                    "fixed_params": fixed_params,
                    "raw_params": raw_params,
                    "body": body,
                }
        except Exception:
            pass

    def on_response(response):
        url = response.url
        try:
            ct = response.headers.get("content-type","")
            if "json" not in ct: return
        except Exception:
            return
        if "vehiclesearch" in url and "/search/" in url:
            try:
                data = response.json()
                t    = (data.get("metadata") or {}).get("totalCount", 0)
                hits = data.get("hits") or []
                with lock:
                    if t and t > total_count[0]: total_count[0] = t
                    for h in hits:
                        hid = hit_id(h)
                        if hid not in seen_ids:
                            seen_ids.add(hid)
                            all_hits.append(h)
                            _v  = h.get("vehicle") or h.get("document") or h.get("data") or h
                            _pr = _v.get("pricing") or (_v.get("vehicleSpecification") or {}).get("pricing") or h.get("pricing") or {}
                            _p  = _pr.get("price") or _pr.get("retailPrice") or _pr.get("listPrice") or {}
                            if _p.get("value") or _p.get("amount") or _p.get("grossAmount"):
                                price_count[0] += 1
                if hits:
                    # ── DIAGNÓSTICO: imprimir claves del primer vehículo (solo la primera vez) ──
                    if os.getenv("BMW_DEBUG_DATE_FIELDS", "0").lower() in ("1", "true", "yes") and len(all_hits) <= len(hits) and hits:
                        def _flat_keys(obj, prefix="", depth=0):
                            if not isinstance(obj, dict) or depth > 5: return []
                            keys = []
                            for k, v in obj.items():
                                full = prefix + k if not prefix else prefix + "." + k
                                keys.append(full)
                                keys.extend(_flat_keys(v, full, depth+1))
                            return keys
                        first_v = hits[0].get("vehicle") or hits[0].get("document") or hits[0].get("data") or hits[0]
                        all_keys = _flat_keys(hits[0])
                        date_keys = [k for k in all_keys if any(x in k.lower() for x in
                                     ["date","since","publish","creat","offer","list","online","launch","avail"])]
                        print("\n\n   ── CAMPOS DE FECHA EN API BMW ──")
                        for dk in sorted(set(date_keys)):
                            # Obtener el valor real
                            parts = dk.split(".")
                            val = hits[0]
                            for p in parts:
                                val = val.get(p, None) if isinstance(val, dict) else None
                            print(f"   {dk}: {val}")
                        print("   ─────────────────────────────────\n")
                    # ─────────────────────────────────────────────────────────────────────────
                    total = total_count[0] or 1
                    pct  = round(len(all_hits)*100/total)
                    ppct = round(price_count[0]*100/len(all_hits)) if all_hits else 0
                    print("\r   Capturados: " + str(len(all_hits)) + "/" + str(total_count[0])
                          + " (" + str(pct) + "%)  | Precios: " + str(price_count[0])
                          + " (" + str(ppct) + "%)", end="", flush=True)
            except Exception:
                pass
        elif "vehiclesearch" in url and "/aggregate/" in url:
            try:
                data = response.json()
                model_ranges = extract_marketing_model_ranges(data)
                if model_ranges:
                    aggregate_model_ranges.append({
                        "total": int(data.get("totalCount") or 0),
                        "model_ranges": model_ranges,
                    })
            except Exception:
                pass

    with sync_playwright() as p:
        print("Abriendo navegador...")
        headless = os.getenv("BMW_HEADLESS", "0").lower() in ("1", "true", "yes")
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
            locale="es-ES", viewport={"width":1440,"height":900},
        )
        page = ctx.new_page()
        # Limpiar estado del browser para evitar filtros guardados
        page.goto("https://www.bmw.es", wait_until="domcontentloaded", timeout=30000)
        try:
            page.evaluate("localStorage.clear(); sessionStorage.clear();")
            ctx.clear_cookies()
        except Exception:
            pass
        page.on("request", on_request)
        page.on("response", on_response)
        page.set_default_timeout(120000)

        print("Cargando BMW Stock Locator...")
        page.goto(START_URL, wait_until="domcontentloaded", timeout=120000)
        accept_cookies(page)
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass

        print("Esperando primera respuesta de la API...")
        for _ in range(80):
            if total_count[0] > 0: break
            page.wait_for_timeout(500)

        if total_count[0] == 0:
            print("   Sin respuesta API inicial; recargando Stock Locator...")
            page.reload(wait_until="domcontentloaded", timeout=120000)
            try:
                page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            for _ in range(120):
                if total_count[0] > 0: break
                page.wait_for_timeout(500)

        if total_count[0] == 0:
            browser.close()
            raise RuntimeError("La pagina no hizo llamadas a la API. Comprueba la conexion.")

        total = total_count[0]
        print("\n   Total inicial: " + str(total))
        page.wait_for_timeout(4000)
        print("   Hits iniciales: " + str(len(all_hits)))
        print("   Request capturado: " + ("SI" if captured_req[0] else "NO"))

        if captured_req[0]:
            req         = captured_req[0]
            base_url    = req["url"]
            fixed_params = req.get("fixed_params", {})
            raw_params  = req.get("raw_params", {})
            real_body   = req.get("body") or {}
            api_headers = dict(req.get("headers") or {})
            api_headers["content-type"] = "application/json"
            api_headers.setdefault("accept", "application/json")
            page_body   = real_body or {}
            # El hash NO es necesario: la API stolo-data-service acepta requests sin hash.
            # maxResults máximo confirmado = 12 (>=24 devuelve 400).
            # Usamos solo brand y context como params fijos.
            base_api_params = {"brand": (raw_params or fixed_params).get("brand", "BMW"),
                               "context": (raw_params or fixed_params).get("context", "results-page")}
            BATCH_SIZE  = int(os.getenv("BMW_BATCH_SIZE", "2") or "2")
            DELAY_MS    = int(os.getenv("BMW_DELAY_MS", "1800") or "1800")
            JITTER_MS   = int(os.getenv("BMW_JITTER_MS", "1200") or "1200")
            MAX_RETRIES = int(os.getenv("BMW_MAX_RETRIES", "2") or "2")
            RATE_LIMIT_DELAY_MS = int(os.getenv("BMW_429_BASE_MS", "75000") or "75000")
            CHECKPOINT_EVERY = int(os.getenv("BMW_CHECKPOINT_EVERY", "240") or "240")
            DEALER_CHUNK_SIZE = int(os.getenv("BMW_DEALER_CHUNK_SIZE", "15") or "15")
            USE_DEALER_CHUNKS = os.getenv("BMW_USE_DEALER_CHUNKS", "0").lower() in ("1", "true", "yes")
            MODEL_PREP_DELAY_MS = int(os.getenv("BMW_MODEL_PREP_DELAY_MS", "700") or "700")
            print("   Request body keys: " + ", ".join(str(k) for k in real_body.keys()))
            print("   Request body preview: " + json.dumps(real_body, ensure_ascii=False)[:800])

            already = len(all_hits)
            first_start = already if already % MAX_PER_PAGE == 0 else (already//MAX_PER_PAGE+1)*MAX_PER_PAGE
            global_remaining = list(range(first_start, total, MAX_PER_PAGE))

            def ingest_hits(hits, total_seen=0):
                added = 0
                with lock:
                    if total_seen and total_seen > total_count[0]:
                        total_count[0] = total_seen
                    for h in hits:
                        hid = hit_id(h)
                        if hid not in seen_ids:
                            seen_ids.add(hid)
                            all_hits.append(h)
                            added += 1
                            _v  = h.get("vehicle") or h.get("document") or h.get("data") or h
                            _pr = _v.get("pricing") or (_v.get("vehicleSpecification") or {}).get("pricing") or h.get("pricing") or {}
                            _p  = _pr.get("price") or _pr.get("retailPrice") or _pr.get("listPrice") or {}
                            if _p.get("value") or _p.get("amount") or _p.get("grossAmount"):
                                price_count[0] += 1
                return added

            def fetch_job(job):
                try:
                    req_p = _urlreq.Request(
                        job["url"],
                        data=json.dumps(job["body"]).encode("utf-8"),
                        headers=api_headers,
                        method="POST",
                    )
                    with _urlreq.urlopen(req_p, timeout=45) as resp_p:
                        data_p = json.loads(resp_p.read().decode("utf-8"))
                    return {
                        "start": job["start"],
                        "key": job.get("key", str(job["start"])),
                        "status": int(getattr(resp_p, "status", 0) or 0),
                        "hits": data_p.get("hits") or [],
                        "totalCount": ((data_p.get("metadata") or {}).get("totalCount") or 0),
                    }
                except _urlerr.HTTPError as e:
                    err_body = ""
                    try:
                        err_body = e.read().decode("utf-8", "replace")[:300]
                    except Exception:
                        pass
                    return {
                        "start": job["start"],
                        "key": job.get("key", str(job["start"])),
                        "status": int(e.code or 0),
                        "error": err_body,
                        "hits": [],
                    }
                except Exception as e:
                    return {
                        "start": job["start"],
                        "key": job.get("key", str(job["start"])),
                        "status": 0,
                        "error": str(e),
                        "hits": [],
                    }

            def wait_with_progress(total_ms, reason):
                total_s = max(1, int(total_ms // 1000))
                print("\n   " + reason + " (" + str(total_s) + "s)")
                remaining_ms = total_ms
                while remaining_ms > 0:
                    step = min(15000, remaining_ms)
                    page.wait_for_timeout(step)
                    remaining_ms -= step
                    if remaining_ms > 0:
                        print("   ... quedan ~" + str(int(remaining_ms // 1000)) + "s", flush=True)

            def chunked(items, size):
                size = max(1, int(size or 1))
                for idx in range(0, len(items), size):
                    yield items[idx:idx + size]

            def body_with_bunos(source_body, dealer_ids):
                body_p = json.loads(json.dumps(source_body))
                sc = body_p.get("searchContext")
                if isinstance(sc, list) and sc and isinstance(sc[0], dict):
                    sc[0]["buNos"] = list(dealer_ids)
                return body_p

            def body_with_marketing_model_range(source_body, model_range):
                body_p = json.loads(json.dumps(source_body))
                sc = body_p.get("searchContext")
                if not isinstance(sc, list) or not sc:
                    body_p["searchContext"] = [{}]
                    sc = body_p["searchContext"]
                if not isinstance(sc[0], dict):
                    sc[0] = {}
                sc[0]["model"] = {"marketingModelRange": {"value": [model_range]}}
                return body_p

            def make_job(start, body_p, key=None):
                qp = dict(base_api_params)
                qp["maxResults"] = str(MAX_PER_PAGE)
                qp["startIndex"] = str(start)
                return {
                    "start": start,
                    "key": key or str(start),
                    "url": base_url + "?" + _up.urlencode(qp),
                    "body": body_p,
                }

            def build_jobs(starts):
                jobs_built = []
                for item in starts:
                    if isinstance(item, dict) and "body" in item and "url" in item:
                        jobs_built.append(item)
                        continue
                    body_p = json.loads(json.dumps(page_body))
                    jobs_built.append(make_job(int(item), body_p))
                return jobs_built

            def run_jobs(jobs_to_run, workers):
                if workers <= 1 or len(jobs_to_run) <= 1:
                    return [fetch_job(job) for job in jobs_to_run]
                with _cf.ThreadPoolExecutor(max_workers=workers) as executor:
                    return list(executor.map(fetch_job, jobs_to_run))

            dealer_ids = []
            search_context = page_body.get("searchContext")
            if isinstance(search_context, list) and search_context and isinstance(search_context[0], dict):
                raw_bunos = search_context[0].get("buNos") or []
                if isinstance(raw_bunos, list):
                    dealer_ids = [str(x) for x in raw_bunos if str(x).strip()]

            pending = []
            model_range_counts = {}
            if aggregate_model_ranges:
                chosen_aggregate = min(
                    aggregate_model_ranges,
                    key=lambda item: abs(int(item.get("total") or 0) - int(total or 0)),
                )
                model_range_counts = dict(chosen_aggregate.get("model_ranges") or {})

            if model_range_counts:
                model_items = sorted(model_range_counts.items(), key=lambda item: (-int(item[1] or 0), item[0]))
                print("   Dividiendo por MARKETING_MODEL_RANGE: " + str(len(model_items))
                      + " grupos, suma facets " + str(sum(int(c or 0) for _, c in model_items)))
                for model_range, expected_count in model_items:
                    model_body = body_with_marketing_model_range(page_body, model_range)
                    first_job = make_job(0, model_body, "mmr:" + model_range + ":0")
                    first_result = fetch_job(first_job)
                    status = int((first_result or {}).get("status") or 0)
                    if status not in range(200, 300):
                        print("   Modelo " + model_range + " fallo en start 0: HTTP " + str(status)
                              + " " + str((first_result or {}).get("error") or "")[:120])
                        pending.append(first_job)
                        continue
                    model_total = int((first_result or {}).get("totalCount") or 0)
                    hits = (first_result or {}).get("hits") or []
                    added = ingest_hits(hits, model_total)
                    print("   Modelo " + model_range + ": total " + str(model_total)
                          + " (facet " + str(expected_count) + "), add " + str(added))
                    if model_total > 1100:
                        print("   Aviso: " + model_range + " supera 1100; puede requerir subparticion.")
                    for start in range(MAX_PER_PAGE, model_total, MAX_PER_PAGE):
                        pending.append(make_job(
                            start,
                            model_body,
                            "mmr:" + model_range + ":" + str(start),
                        ))
                    if MODEL_PREP_DELAY_MS:
                        page.wait_for_timeout(MODEL_PREP_DELAY_MS)
            elif dealer_ids and USE_DEALER_CHUNKS:
                dealer_chunks = list(chunked(dealer_ids, DEALER_CHUNK_SIZE))
                print("   Dividiendo " + str(len(dealer_ids)) + " concesionarios en "
                      + str(len(dealer_chunks)) + " grupos de max " + str(DEALER_CHUNK_SIZE))
                for chunk_i, dealer_chunk in enumerate(dealer_chunks, 1):
                    chunk_body = body_with_bunos(page_body, dealer_chunk)
                    first_job = make_job(0, chunk_body, "g" + str(chunk_i) + ":0")
                    first_result = fetch_job(first_job)
                    status = int((first_result or {}).get("status") or 0)
                    if status not in range(200, 300):
                        print("   Grupo " + str(chunk_i) + " fallo en start 0: HTTP " + str(status)
                              + " " + str((first_result or {}).get("error") or "")[:120])
                        pending.append(first_job)
                        continue
                    chunk_total = int((first_result or {}).get("totalCount") or 0)
                    hits = (first_result or {}).get("hits") or []
                    added = ingest_hits(hits, chunk_total)
                    print("   Grupo " + str(chunk_i) + "/" + str(len(dealer_chunks))
                          + ": " + str(len(dealer_chunk)) + " dealers, total "
                          + str(chunk_total) + ", add " + str(added))
                    if chunk_total > 1100:
                        print("   Aviso: grupo " + str(chunk_i) + " tiene " + str(chunk_total)
                              + " resultados; baja BMW_DEALER_CHUNK_SIZE si aparece 400.")
                    for start in range(MAX_PER_PAGE, chunk_total, MAX_PER_PAGE):
                        pending.append(make_job(start, chunk_body, "g" + str(chunk_i) + ":" + str(start)))
            else:
                pending = [make_job(start, page_body, str(start)) for start in global_remaining]

            print("   Paginas pendientes: " + str(len(pending)))
            print("   Paginando en batches de " + str(BATCH_SIZE) + " (particionado, con pricing)...")
            retry_rounds = {}
            rate_limit_mode = False
            last_checkpoint = len(all_hits)

            while pending:
                if MAX_VEHICLES and len(all_hits) >= MAX_VEHICLES: break
                current_batch_size = 1 if rate_limit_mode else max(1, BATCH_SIZE)
                batch = [pending.pop(0) for _ in range(min(current_batch_size, len(pending)))]
                jobs = build_jobs(batch)
                results = run_jobs(jobs, current_batch_size)

                # Reintentar jobs fallidos con backoff en 429
                for attempt in range(MAX_RETRIES):
                    retryable_jobs = []
                    for job, result in zip(jobs, results):
                        status = int((result or {}).get("status") or 0)
                        if status in (0, 408, 425, 429, 500, 502, 503, 504):
                            retryable_jobs.append(job)
                    if not retryable_jobs:
                        break
                    is_429 = any(int((r or {}).get("status") or 0) == 429
                                 for r in results if isinstance(r, dict))
                    if is_429:
                        rate_limit_mode = True
                    wait_ms = RATE_LIMIT_DELAY_MS if is_429 else 5000 * (attempt + 1)
                    if is_429:
                        print("\n   HTTP 429 — enfriando " + str(wait_ms // 1000) + "s...")
                    if is_429:
                        wait_with_progress(wait_ms, "HTTP 429 - enfriando y bajando a modo secuencial")
                    else:
                        wait_with_progress(wait_ms, "Error temporal - reintentando")
                    retry_results = run_jobs(retryable_jobs, 1)
                    retry_by_key = {r.get("key"): r for r in retry_results}
                    results = [retry_by_key.get(r.get("key"), r) for r in results]

                batch_hits = 0
                batch_added = 0
                empty_offsets = []
                failed_offsets = []
                job_by_key = {str(job.get("key", job.get("start"))): job for job in jobs}
                for result in results:
                    if not isinstance(result, dict):
                        continue
                    status = int(result.get("status") or 0)
                    start = result.get("start")
                    key = str(result.get("key") or start)
                    hits = result.get("hits") or []
                    if status not in range(200, 300):
                        failed_offsets.append(key + ":" + str(status))
                    elif not hits:
                        empty_offsets.append(key)
                    batch_hits += len(hits)
                    batch_added += ingest_hits(hits, int(result.get("totalCount") or 0))

                if failed_offsets:
                    print("\n   Errores persistentes: " + ", ".join(failed_offsets[:8]))
                    hard_fail = False
                    for item in failed_offsets:
                        parts = item.split(":")
                        if len(parts) < 2:
                            continue
                        status_s = parts[-1]
                        key = ":".join(parts[:-1])
                        if status_s not in ("0", "408", "425", "429", "500", "502", "503", "504"):
                            hard_fail = True
                            continue
                        retry_rounds[key] = retry_rounds.get(key, 0) + 1
                        if retry_rounds[key] <= 3 and key in job_by_key:
                            pending.append(job_by_key[key])
                    if any(":429" in item for item in failed_offsets):
                        rate_limit_mode = True
                    if hard_fail:
                        print("\n   Corte de paginacion: BMW devuelve HTTP no recuperable. "
                              "Se guardara el stock parcial capturado.")
                        break
                elif empty_offsets and batch_hits == 0:
                    print("\n   Batch sin hits en offsets: " + ", ".join(empty_offsets[:8]) + " (continua)")

                pct  = round(len(all_hits)*100/total_count[0]) if total_count[0] else 0
                ppct = round(price_count[0]*100/len(all_hits)) if all_hits else 0
                print("\r   Capturados: " + str(len(all_hits)) + "/" + str(total_count[0])
                      + " (" + str(pct) + "%)  | Precios: " + str(price_count[0])
                      + " (" + str(ppct) + "%)  | Batch hits/add: "
                      + str(batch_hits) + "/" + str(batch_added), end="", flush=True)
                if CHECKPOINT_EVERY and len(all_hits) - last_checkpoint >= CHECKPOINT_EVERY:
                    try:
                        partial_rows = [map_vehicle(h, pricing_map) for h in all_hits]
                        save_excel([r for r in partial_rows if r], PARTIAL_XLSX)
                        last_checkpoint = len(all_hits)
                        print("\n   Checkpoint guardado: " + str(Path(PARTIAL_XLSX).resolve())
                              + " (" + str(last_checkpoint) + " vehiculos)")
                    except Exception as e:
                        print("\n   Aviso: no se pudo guardar checkpoint parcial: " + str(e))
                delay_ms = DELAY_MS + random.randint(0, max(0, JITTER_MS))
                if rate_limit_mode:
                    delay_ms = max(delay_ms, 8000 + random.randint(0, 5000))
                page.wait_for_timeout(delay_ms)
        else:
            print("   AVISO: Request no capturado.")

        print("\n   Parseando " + str(len(all_hits)) + " vehiculos...")
        browser.close()

    if pricing_map:
        print("   Pricing map: " + str(len(pricing_map)))
    rows = [map_vehicle(h, pricing_map) for h in all_hits]
    rows = [r for r in rows if r]
    unique_rows = []
    seen_car_ids = set()
    for row in rows:
        car_id = row.get("Car ID") or ""
        if car_id and car_id in seen_car_ids:
            continue
        if car_id:
            seen_car_ids.add(car_id)
        unique_rows.append(row)
    rows = unique_rows

    from collections import Counter
    by_model  = Counter(r.get("Carline") or "SIN MODELO" for r in rows)
    sin_precio = Counter(r.get("Carline") or "SIN MODELO" for r in rows if not r.get("PVP (EUR)"))
    print("\n  --- MODELOS capturados ---")
    print("  {:<28s} {:>6s}  {:>10s}".format("Carline","Total","Sin precio"))
    for m, cnt in by_model.most_common():
        sp = sin_precio.get(m, 0)
        flag = " !!!" if sp == cnt else ""
        print("  {:<28s} {:>6d}  {:>10d}{}".format(m[:28], cnt, sp, flag))
    return rows


def main():
    sep = "=" * 65
    start = datetime.now()
    print(sep)
    print("  BMW ESPANA -- STOCK SCRAPER")
    print("  " + start.strftime("%Y-%m-%d %H:%M:%S"))
    print("  Source: bmw.es Stock Locator")
    print(sep)
    rows = scrape_bmw()
    elapsed = datetime.now() - start
    save_excel(rows, OUT_XLSX)
    n  = len(rows)
    wp = sum(1 for r in rows if r.get("PVP (EUR)"))
    wc = sum(1 for r in rows if r.get("Ciudad"))
    print("")
    print(sep)
    print("  RESUMEN")
    print("  Vehiculos  : " + str(n))
    print("  Con precio : " + str(wp) + " (" + str(round(wp*100/max(n,1))) + "%)")
    print("  Con ciudad : " + str(wc) + " (" + str(round(wc*100/max(n,1))) + "%)")
    print("  Excel      : " + str(Path(OUT_XLSX).resolve()))
    print("  Tiempo     : " + str(elapsed))
    print(sep)


if __name__ == "__main__":
    import traceback
    try:
        main()
    except Exception:
        print("\n*** ERROR EN EL SCRAPER ***")
        traceback.print_exc()
        raise SystemExit(1)
