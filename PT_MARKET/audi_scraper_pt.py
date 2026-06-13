"""
audi_scraper_pt.py
==================
Scraper for ALL Audi vehicles in stock in Portugal.
Source: disponivel-imediatamente.audi.pt (Porsche Informatik marketplace, SSR HTML)

Strategy: The site uses server-side rendering. The landing page lists all available
model groups (e.g. ?mg=Audi+Q5). We discover them dynamically on every run, then
fetch each group (up to 100 cars per page) to collect the full catalogue.
This means new models added by Audi PT are picked up automatically.

Install dependencies:
    pip install requests beautifulsoup4 openpyxl

Usage:
    python audi_scraper_pt.py

Output:
    STOCK_AUDI_PT.xlsx
"""

import requests
import re
import time
import os
import sys
from datetime import datetime
from urllib.parse import quote_plus, unquote_plus
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ═══════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════

BASE_URL     = "https://disponivel-imediatamente.audi.pt"
OUT_XLSX     = "STOCK_AUDI_PT.xlsx"
PAGE_SIZE    = 100     # max reliable page size
DELAY        = 0.4     # seconds between requests
FETCH_DETAIL = True    # fetch individual car pages for version/city/detail enrichment

# Fallback list used ONLY if the landing page cannot be scraped
MODEL_GROUPS_FALLBACK = [
    "Audi A1 Sportback", "Audi A1 allstreet",
    "Audi A3 Limousine", "Audi A3 Sportback", "Audi A3 allstreet",
    "Audi A5 Avant", "Audi A5 Limousine",
    "Audi A6 Avant", "Audi A6 Avant e-tron", "Audi A6 Limousine", "Audi A6 Sportback e-tron",
    "Audi Q2",
    "Audi Q3 SUV", "Audi Q3 Sportback",
    "Audi Q4 SUV e-tron", "Audi Q4 Sportback e-tron",
    "Audi Q5", "Audi Q5 Sportback",
    "Audi Q6 SUV e-tron", "Audi Q6 Sportback e-tron",
]


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\u200d", "").strip()
    if "Ã" in text or "Â" in text:
        try:
            text = text.encode("latin1").decode("utf-8")
        except Exception:
            pass
    return re.sub(r"\s+", " ", text).strip()


def parse_eur(value):
    text = clean_text(value)
    text = text.replace("€", "").replace(" ", "").replace("-", "").strip()
    if not text:
        return ""
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return ""


# ═══════════════════════════════════════════════════════
# HTTP SESSION
# ═══════════════════════════════════════════════════════

def normalize_date_text(value):
    text = clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text[:10] if re.match(r"\d{4}-\d{2}-\d{2}", text) else ""


def extract_publication_date(text):
    text = clean_text(text)
    if not text:
        return ""
    date_re = r"(\d{4}-\d{2}-\d{2}|\d{1,2}[./-]\d{1,2}[./-]\d{2,4})"
    for line in re.split(r"[\n|]", text):
        key = line.lower()
        if any(token in key for token in ("publicad", "adicionad", "online", "desde", "data")):
            match = re.search(date_re, line)
            if match:
                return normalize_date_text(match.group(1))
    return ""

def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
        "Referer": BASE_URL,
    })
    return s


# ═══════════════════════════════════════════════════════
# DYNAMIC MODEL GROUP DISCOVERY
# ═══════════════════════════════════════════════════════

def discover_model_groups(session):
    """
    Scrape the landing page to get all current model group links.
    Returns a list of decoded model group strings, e.g. ['Audi Q5', 'Audi A3 Sportback', ...]
    Falls back to MODEL_GROUPS_FALLBACK if the page can't be reached.
    """
    try:
        r = session.get(BASE_URL + "/", timeout=20)
        r.encoding = "utf-8"
        r.raise_for_status()
        # Model group links look like: href="/search?mg=Audi+Q5"
        raw_mgs = re.findall(r'/search\?mg=([^"&\s]+)', r.text)
        groups = sorted(set(unquote_plus(mg).replace("+", " ") for mg in raw_mgs))
        if groups:
            print(f"  Discovered {len(groups)} model groups from landing page")
            return groups
        else:
            print("  Warning: no model groups found on landing page - using fallback list")
            return MODEL_GROUPS_FALLBACK
    except Exception as e:
        print(f"  Warning: could not reach landing page ({e}) - using fallback list")
        return MODEL_GROUPS_FALLBACK


# ═══════════════════════════════════════════════════════
# CITY INFERENCE FROM DEALER NAME
# ═══════════════════════════════════════════════════════

# Portuguese cities / towns commonly found in dealer names
_PT_CITIES = [
    "Lisboa", "Porto", "Braga", "Coimbra", "Aveiro", "Faro", "Évora", "Evora",
    "Setúbal", "Setubal", "Viseu", "Leiria", "Santarém", "Santarem",
    "Castelo Branco", "Guarda", "Bragança", "Braganca", "Vila Real",
    "Viana do Castelo", "Funchal", "Loures", "Cascais", "Sintra", "Almada",
    "Oeiras", "Amadora", "Matosinhos", "Gaia", "Gondomar",
    "Guimarães", "Guimaraes", "Barcelos", "Famalicão", "Famalicao",
    "Braga", "Maia", "Valongo", "Paredes", "Penafiel",
    "Loulé", "Loule", "Portimão", "Portimao", "Albufeira", "Olhão", "Olhao",
    "Tavira", "Lagos", "Silves", "Vila Real de Santo António",
    "Lamego", "Chaves", "Mirandela", "Peso da Régua",
    "Caldas da Rainha", "Tomar", "Torres Vedras", "Entroncamento",
    "Setúbal", "Palmela", "Barreiro", "Montijo", "Alcochete",
    "Seixal", "Sesimbra", "Santiago do Cacém", "Beja", "Elvas", "Portalegre",
    "Estremoz", "Vendas Novas", "Moura", "Serpa",
    "Peniche", "Nazaré", "Alcobaça", "Batalha", "Fátima", "Ourém",
    "Covilhã", "Covilha", "Fundão", "Fundao", "Seia",
    "Vidago", "Mirandela", "Macedo de Cavaleiros",
    "Vouzela", "São João da Madeira", "Santa Maria da Feira",
    "Espinho", "Póvoa de Varzim", "Povoa de Varzim", "Vila do Conde",
    "Felgueiras", "Fafe", "Paços de Ferreira",
    "Coimbra", "Cantanhede", "Figueira da Foz", "Oliveira do Hospital",
    "Mortágua", "Mealhada",
]
# Sort longest first so "Vila Real de Santo António" matches before "Vila Real"
_PT_CITIES_SORTED = sorted(_PT_CITIES, key=len, reverse=True)

def _city_from_dealer(dealer_name):
    """Try to extract a Portuguese city from a dealer name string."""
    if not dealer_name:
        return ""
    dl = dealer_name
    for city in _PT_CITIES_SORTED:
        # case-insensitive word-boundary match
        if re.search(r'(?i)\b' + re.escape(city) + r'\b', dl):
            return city
    return ""


# ═══════════════════════════════════════════════════════
# PARSING — SEARCH LIST PAGE
# ═══════════════════════════════════════════════════════

def parse_search_page(html, model_group):
    """Extract all car cards from a model-group search page."""
    soup = BeautifulSoup(html, "html.parser")

    # Each car is wrapped in a div with class 'group/car-preview'
    car_links = soup.find_all("a", href=re.compile(r"/search/car/"))
    seen  = set()
    cars  = []

    for link in car_links:
        href = link.get("href", "")
        if href in seen:
            continue
        seen.add(href)

        car_id = href.split("/search/car/")[-1]
        div    = link.find_parent("div", class_=lambda c: c and "group/car-preview" in c)
        if not div:
            continue

        lines = [
            clean_text(p)
            for p in div.get_text(separator="\n", strip=True).split("\n")
            if clean_text(p)
        ]
        text = clean_text(div.get_text(separator=" ", strip=True))

        year_idx = next((i for i, p in enumerate(lines) if re.search(r"MY\s*\d{4}", p)), None)
        model = ""
        if year_idx is not None:
            for p in reversed(lines[:year_idx]):
                if p not in {"Disponível Imediatamente", "Carregando imagem"}:
                    model = p
                    break

        # Year from "Veículos novos - MY 2025"
        year_m = re.search(r"MY\s*(\d{4})", text)
        year   = year_m.group(1) if year_m else ""

        # Fuel
        fuel_keywords = {"Gasolina", "Diesel", "Elétrico", "Plug-In-Hibrido", "Plug-In-Híbrido", "Hibrido", "Híbrido", "GPL"}
        fuel = next((p for p in lines if p in fuel_keywords), "")
        if not fuel and "e-tron" in f"{model} {model_group}".lower():
            fuel = "Elétrico"
        if not fuel and re.search(r"\bTFSIe\b", model, re.IGNORECASE):
            fuel = "Plug-In-Híbrido"
        if not fuel and "e-hybrid" in model.lower():
            fuel = "Plug-In-Híbrido"
        if not fuel and re.search(r"\bTFSI\b", model, re.IGNORECASE):
            fuel = "Gasolina"
        if not fuel and re.search(r"\bTDI\b", model, re.IGNORECASE):
            fuel = "Diesel"

        # Power: look for "95 cv / 70 kW" pattern or separate tokens
        pwr_m = re.search(r"(\d+)\s*cv\s*/\s*(\d+)\s*kW", text, re.IGNORECASE)
        if pwr_m:
            power = f"{pwr_m.group(2)} kW ({pwr_m.group(1)} HP)"
        else:
            # Try individual kW
            kw_m = re.search(r"(\d+)\s*kW", text)
            power = f"{kw_m.group(1)} kW" if kw_m else ""

        # Body type
        body_m = re.search(r"(\d+)\s*portas", text)
        body   = f"{body_m.group(1)}-door" if body_m else ""

        # Drive (may appear in some cards)
        drive_m = re.search(r"(Tração\s+\w+(?:\s+\w+)?)", text, re.IGNORECASE)
        drive  = drive_m.group(1) if drive_m else ""

        # Color follows the body line in the SSR card.
        ext_color = ""
        body_idx = next((i for i, p in enumerate(lines) if re.search(r"\d+\s*portas", p)), None)
        if body_idx is not None and body_idx + 1 < len(lines):
            ext_color = lines[body_idx + 1]

        # Dealer name (after "Disponível em")
        dealer = ""
        dealer_idx = next((i for i, p in enumerate(lines) if p == "Disponível em"), None)
        if dealer_idx is not None and dealer_idx + 1 < len(lines):
            dealer = lines[dealer_idx + 1]

        # City — try to extract from dealer name (e.g. "Jomecsport Braga" → "Braga")
        city = _city_from_dealer(dealer)

        # Prices - first is PVPR with taxes, second is base price
        price_lines = [p for p in lines if "€" in p and re.search(r"\d", p)]
        rrp = parse_eur(price_lines[0]) if price_lines else ""
        base_price = parse_eur(price_lines[1]) if len(price_lines) > 1 else ""

        # Availability
        if "Disponível Imediatamente" in text or "Disponível imediatamente" in text:
            availability = "available"
        else:
            availability = "soon"
        published_date = extract_publication_date(text)

        cars.append({
            "car_id":       car_id,
            "model_group":  model_group,
            "model":        model,
            "version":      model,
            "year":         year,
            "fuel":         fuel,
            "power":        power,
            "body":         body,
            "drive":        drive,
            "ext_color":    ext_color,
            "dealer":       dealer,
            "city":         city,
            "rrp_eur":      rrp,
            "base_price":   base_price,
            "availability": availability,
            "published_date": published_date,
            "url":          f"{BASE_URL}/search/car/{car_id}",
        })

    return cars


# ═══════════════════════════════════════════════════════
# PARSING — CAR DETAIL PAGE (optional)
# ═══════════════════════════════════════════════════════

def parse_detail_page(html, car_id):
    """Extract extra fields from individual car detail page."""
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text(separator="\n", strip=True)
    lines = [l.strip() for l in text.split("\n") if l.strip()]

    h1 = soup.find("h1")
    version = clean_text(h1.get_text(" ", strip=True)) if h1 else ""

    dealer = ""
    dealer_link = soup.find("a", href=re.compile(r"/opening-hours/"))
    if dealer_link:
        dealer = clean_text(dealer_link.get_text(" ", strip=True))

    city = ""
    for line in lines:
        line_clean = clean_text(line)
        for candidate in _PT_CITIES_SORTED:
            candidate_clean = clean_text(candidate)
            if re.fullmatch(re.escape(candidate_clean), line_clean, re.IGNORECASE):
                city = candidate_clean
                break
            if re.search(r"(?i)\b" + re.escape(candidate_clean) + r"\b", line_clean):
                city = candidate_clean
                break
        if city:
            break

    # Engine line: "Gasolina | 70 kW (95 cv) | 1,0 l"
    engine_m = re.search(r"(Gasolina|Diesel|Elétrico|Híbrido|Plug-In-Hibrido)[^\n]*\|?\s*(\d+\s*kW.*?\|.*?l)", text)

    # Ext color
    ext_color = ""
    int_color = ""
    for i, line in enumerate(lines):
        if "Cor exterior" in line and i + 1 < len(lines):
            ext_color = lines[i + 1]
        if "Cor Interior" in line and i + 1 < len(lines):
            int_color = lines[i + 1]

    # Dealer address & phone
    dealer_phone = ""
    phone_m = re.search(r"\+351\s*[\d\s]+", text)
    if phone_m:
        dealer_phone = phone_m.group().strip()

    # CO2 emissions
    co2_m = re.search(r"(\d+)\s*g/km", text)
    co2   = co2_m.group(1) if co2_m else ""

    # Consumption
    cons_m = re.search(r"([\d,]+)\s*l/100km", text)
    consumption = cons_m.group(1) if cons_m else ""

    return {
        "version":     version,
        "dealer":      dealer,
        "city":        city,
        "ext_color":   ext_color,
        "int_color":   int_color,
        "dealer_phone": dealer_phone,
        "co2":         co2,
        "consumption": consumption,
        "published_date": extract_publication_date(text),
    }


# ═══════════════════════════════════════════════════════
# SCRAPING
# ═══════════════════════════════════════════════════════

def fetch_model_group(session, model_group):
    """Fetch all cars for a given model group."""
    mg_encoded = quote_plus(model_group)
    url = f"{BASE_URL}/search?mg={mg_encoded}&size={PAGE_SIZE}&sort=FIELD%3APRICE%3AASC"
    try:
        r = session.get(url, timeout=25)
        r.encoding = "utf-8"
        r.raise_for_status()
        return parse_search_page(r.text, model_group)
    except Exception as e:
        print(f"    Warning: Error fetching {model_group}: {e}")
        return []


def fetch_detail(session, car):
    """Fetch detail page and enrich car dict in-place."""
    try:
        r = session.get(car["url"], timeout=20)
        r.encoding = "utf-8"
        if r.status_code == 200:
            extra = parse_detail_page(r.text, car["car_id"])
            # Only override if detail page has richer data
            if extra.get("version"):
                car["version"] = extra["version"]
            if extra.get("dealer") and not car.get("dealer"):
                car["dealer"] = extra["dealer"]
            if extra.get("city") and not car.get("city"):
                car["city"] = extra["city"]
            if extra.get("ext_color") and not car.get("ext_color"):
                car["ext_color"] = extra["ext_color"]
            if extra.get("published_date") and not car.get("published_date"):
                car["published_date"] = extra["published_date"]
            car["int_color"]    = extra.get("int_color", "")
            car["dealer_phone"] = extra.get("dealer_phone", "")
            car["co2"]          = extra.get("co2", "")
            car["consumption"]  = extra.get("consumption", "")
    except Exception:
        pass


# ═══════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════

HEADERS = [
    "Type", "Model", "Version", "Model Group", "Year", "Fuel", "Power", "Body", "Drive",
    "Ext. Color", "Int. Color",
    "RRP (EUR)", "Base Price (EUR)",
    "Availability", "Published Date", "Dealer", "City", "Car ID", "Car URL",
]
FIELD_MAP = [
    "type", "model", "version", "model_group", "year", "fuel", "power", "body", "drive",
    "ext_color", "int_color",
    "rrp_eur", "base_price",
    "availability", "published_date", "dealer", "city", "car_id", "url",
]
WIDTHS = [8,40,40,22,6,18,14,10,18,20,16,14,14,14,18,35,22,22,65]
AUDI_GRAY  = "4E4E4E"
AUDI_LIGHT = "F0F0F0"


def save_excel(cars, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cars in stock"
    ws.freeze_panes = "A2"
    hdr_fill = PatternFill("solid", fgColor=AUDI_GRAY)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor=AUDI_LIGHT)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row_i, car in enumerate(cars, 2):
        for col_i, field in enumerate(FIELD_MAP, 1):
            val = car.get(field, "")
            if val == "" and field in ("rrp_eur", "base_price"): val = None
            if isinstance(val, (dict, list)): val = str(val)
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if row_i % 2 == 0: cell.fill = alt_fill
    for col_i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = w
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary by Model")
    ws2.append(["Model Group", "Vehicles in stock"])
    cnt = Counter(c.get("model_group", "?") for c in cars)
    for mg, count in sorted(cnt.items(), key=lambda x: -x[1]):
        ws2.append([mg, count])
    for col in ws2.iter_cols(min_row=1, max_row=1):
        for cell in col: cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30; ws2.column_dimensions["B"].width = 20

    ws3 = wb.create_sheet("Summary by Dealer")
    ws3.append(["Dealer", "City", "Vehicles in stock"])
    cnt3 = Counter((c.get("dealer", "?"), c.get("city", "")) for c in cars)
    for (dealer, city), count in sorted(cnt3.items(), key=lambda x: -x[1]):
        ws3.append([dealer, city, count])
    for col in ws3.iter_cols(min_row=1, max_row=1):
        for cell in col: cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 40; ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 20

    wb.save(path)
    print(f"  ✓ Excel saved: {path}")


def main():
    start = datetime.now()
    print("=" * 65)
    print("  AUDI PORTUGAL — STOCK SCRAPER")
    print(f"  {start.strftime('%Y-%m-%d %H:%M:%S')}")
    print("  Source: disponivel-imediatamente.audi.pt")
    print("=" * 65)

    session  = make_session()
    all_cars = []
    seen_ids = set()

    print("\n[1/2] Discovering model groups...")
    model_groups = discover_model_groups(session)
    time.sleep(DELAY)

    print(f"\n      Fetching {len(model_groups)} model groups...")
    for i, mg in enumerate(model_groups, 1):
        cars = fetch_model_group(session, mg)
        new_cars = []
        for c in cars:
            if c["car_id"] not in seen_ids:
                seen_ids.add(c["car_id"])
                c["type"] = "New"
                new_cars.append(c)
        all_cars.extend(new_cars)
        print(f"  [{i:02d}/{len(model_groups)}] {mg:<30} -> {len(new_cars)} cars (total: {len(all_cars)})")
        time.sleep(DELAY)

    print(f"\n  -> {len(all_cars)} unique vehicles collected")

    if FETCH_DETAIL:
        print(f"\n[2/2] Fetching detail pages for {len(all_cars)} cars...")
        for i, car in enumerate(all_cars, 1):
            fetch_detail(session, car)
            if i % 20 == 0:
                print(f"  {i}/{len(all_cars)} detail pages fetched")
            time.sleep(DELAY)
    else:
        print("\n[2/2] Skipping detail pages (FETCH_DETAIL=False)")

    elapsed = datetime.now() - start
    print(f"\nSaving Excel...")
    save_excel(all_cars, OUT_XLSX)

    print("\n" + "=" * 65)
    print("  FINAL SUMMARY")
    print(f"  Total vehicles     : {len(all_cars)}")
    city_cnt = sum(1 for c in all_cars if c.get("city"))
    print(f"  With city inferred : {city_cnt}/{len(all_cars)}")
    for mg, n in Counter(c.get("model_group","?") for c in all_cars).most_common(8):
        print(f"    {mg:<35} {n:>4} units")
    print(f"\n  Excel file : {os.path.abspath(OUT_XLSX)}")
    print(f"  Total time : {elapsed}")
    print("=" * 65)


if __name__ == "__main__":
    main()
