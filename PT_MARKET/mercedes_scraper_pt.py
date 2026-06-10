"""
mercedes_scraper_pt.py
======================
Scraper for Mercedes-Benz new cars in stock in Portugal.
Source: eu.api.oneweb.mercedes-benz.com/commerce/onesearch/graphql

Strategy:
  1. Playwright loads the search page and captures the GraphQL request
     (URL + auth headers + query body with variables).
  2. We replay that request directly, incrementing `page` from 0 to N,
     until we have all vehicles (total = 1361 as of June 2026).
  3. Extraction specifically handles the OnewWeb GraphQL response:
     data > any_key > hits[] array.

Install:
    pip install openpyxl playwright
    python -m playwright install chromium

Usage:
    python mercedes_scraper_pt.py  ->  STOCK_MERCEDES_PT.xlsx
"""

from __future__ import annotations
import asyncio, json, re, os, queue, time, sys
from datetime import datetime
from pathlib import Path
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ═══════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════

START_URL = (
    "https://www.mercedes-benz.pt/passengercars/buy/new-car/"
    "search-results.html/vehicleCategory-new-passenger-cars/sortType-price-asc"
)
OUT_XLSX     = "STOCK_MERCEDES_PT.xlsx"
PAGE_SIZE    = 12    # MB default limit per GraphQL page
CAPTURE_WAIT = 40   # seconds to wait for GraphQL request to be captured

# ═══════════════════════════════════════════════════════
# COLUMN DEFINITIONS
# ═══════════════════════════════════════════════════════

HEADERS = [
    "Type", "Model", "Version", "Body", "Fuel", "Gearbox", "Power", "Drive",
    "Ext. Color", "Int. Color",
    "RRP (EUR)", "Monthly Rate (EUR)", "APR (%)",
    "Availability", "Published Date", "Dealer", "City", "Car ID", "Car URL",
]
FIELD_MAP = [
    "type", "model", "version", "body", "fuel", "gearbox", "power", "drive",
    "ext_color", "int_color",
    "rrp_eur", "monthly_rate", "apr",
    "availability", "published_date", "dealer", "city", "car_id", "url",
]
WIDTHS = [8, 30, 40, 14, 14, 12, 12, 14, 20, 16, 14, 16, 8, 14, 18, 35, 18, 22, 65]
MB_BLUE  = "00ADEF"
MB_LIGHT = "EAF6FF"


# ═══════════════════════════════════════════════════════
# GRAPHQL RESPONSE PARSING
# ═══════════════════════════════════════════════════════

def extract_hits(data):
    """
    Extract (hits_list, total_count) from a OnewWeb GraphQL response.
    GraphQL always wraps everything under "data", then nested keys.
    We look for the first array named hits / nodes / edges / vehicles / results.
    """
    if not isinstance(data, dict):
        return [], 0

    # Unwrap GraphQL envelope
    root = data.get("data") or data

    def _search(obj, depth=0):
        if depth > 6 or not isinstance(obj, dict):
            return [], 0
        for key in ("hits", "nodes", "edges", "vehicles", "results",
                    "items", "content", "list"):
            val = obj.get(key)
            if isinstance(val, list) and val and isinstance(val[0], dict):
                total = (obj.get("count") or obj.get("total")
                         or obj.get("totalCount") or obj.get("totalHits")
                         or _s(obj, "navigation", "totalResults") or 0)
                return val, int(total)
        for v in obj.values():
            if isinstance(v, dict):
                hits, total = _search(v, depth + 1)
                if hits:
                    return hits, total
        return [], 0

    return _search(root)


# ═══════════════════════════════════════════════════════
# FIELD EXTRACTION FROM A SINGLE VEHICLE HIT
# ═══════════════════════════════════════════════════════

def _s(obj, *keys, default=""):
    for k in keys:
        if obj is None: return default
        if isinstance(obj, dict): obj = obj.get(k)
        elif isinstance(obj, list):
            try: obj = obj[int(k)]
            except: return default
        else: return default
    return obj if obj is not None else default

def _str(v):
    if not v: return ""
    if isinstance(v, dict):
        return (v.get("formattedValue") or v.get("description") or v.get("name") or v.get("label")
                or v.get("value") or v.get("text") or "")
    if isinstance(v, list) and v:
        return _str(v[0])
    return str(v)

def _num(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, dict):
        for k in ("grossAmount","amount","value","gross","net"):
            if v.get(k) is not None:
                return _num(v[k])
        return None
    try:
        return float(str(v).replace(" ", "").replace(".", "").replace(",", "."))
    except: return None

FUEL_MAP  = {"PETROL":"Petrol","GASOLINE":"Petrol","DIESEL":"Diesel",
             "ELECTRIC":"Electric","BEV":"Electric","PHEV":"Plug-in Hybrid",
             "HYBRID":"Hybrid","MHEV":"Mild Hybrid","HEV":"Hybrid"}
TRANS_MAP = {"AUTOMATIC":"Automatic","MANUAL":"Manual",
             "9G_TRONIC":"9G-TRONIC","7G_TRONIC":"7G-TRONIC"}
DRIVE_MAP = {"RWD":"Rear-wheel","FWD":"Front-wheel",
             "AWD":"4MATIC AWD","4MATIC":"4MATIC AWD"}

PRICE_API = "https://eu.api.oneweb.mercedes-benz.com/commerce/pls/public/v6/ncos/markets/PT/prices/{car_id}"


def extract_vehicle(hit):
    """
    Extract standardised vehicle record from a MB onesearch GraphQL hit.
    The hit may be the vehicle itself or a wrapper with a nested 'vehicle' key.
    We try both shapes.
    """
    # Some hits wrap vehicle data one level deeper
    v = hit.get("vehicle") or hit.get("vehicleData") or hit

    # Model name - try many common MB API field names
    model = (_str(_s(v,"modelSeries","description"))
             or _str(_s(v,"classDescription"))
             or _str(_s(v,"model","description"))
             or _str(_s(v,"modelDescription"))
             or _str(_s(v,"nameplate","modelDescription"))
             or _str(_s(v,"vehicleModel","vehicleClass"))
             or _str(_s(v,"name"))
             or _str(_s(v,"displayName"))
             or _str(_s(hit,"modelSeries","description"))
             or _str(_s(hit,"classDescription"))
             or "")

    version = (_str(_s(v,"variantDescription"))
               or _str(_s(v,"equipmentLine","description"))
               or _str(_s(v,"derivative"))
               or _str(_s(v,"subtitle"))
               or _str(_s(v,"line","description"))
               or _str(_s(v,"vehicleModel","motorization"))
               or _str(_s(v,"vehicleModel","name"))
               or "")

    body = (_str(_s(v,"bodyType","description"))
            or _str(_s(v,"vehicleModel","bodyType"))
            or _str(_s(v,"bodyStyle"))
            or _str(_s(v,"category"))
            or "")

    raw_fuel = (_str(_s(v,"fuel","description"))
                or _str(_s(v,"technicalInformation","engine","fuelType"))
                or _str(_s(v,"fuelType"))
                or _str(_s(v,"engineType"))
                or _str(_s(v,"powertrainType"))
                or "")
    fuel = FUEL_MAP.get(raw_fuel.upper(), raw_fuel)

    raw_gear = _str(_s(v,"transmission","description")
                    or _s(v,"technicalInformation","transmission")
                    or _s(v,"gearbox"))
    gearbox  = TRANS_MAP.get(raw_gear.upper(), raw_gear)

    kw = (_s(v,"powerKW") or _s(v,"enginePower","kw") or _s(v,"power","kw"))
    hp = (_s(v,"powerHP") or _s(v,"enginePower","hp") or _s(v,"power","hp"))
    power = (_str(_s(v,"technicalInformation","engine","power"))
             or (f"{kw} kW ({hp} HP)" if kw and hp else (f"{kw} kW" if kw else "")))

    raw_drive = _str(_s(v,"driveType") or _s(v,"drive") or _s(v,"wheelDrive"))
    drive = DRIVE_MAP.get(raw_drive.upper(), raw_drive)

    ext_color = (_str(_s(v,"exteriorColor","description"))
                 or _str(_s(v,"color","description"))
                 or _str(_s(v,"paintDescription"))
                 or _str(_s(v,"colorName"))
                 or "")
    int_color = (_str(_s(v,"interiorColor","description"))
                 or _str(_s(v,"upholsteryColor"))
                 or _str(_s(v,"interior","description"))
                 or "")

    # Price: try several price fields
    price_obj = (_s(v,"price") or _s(v,"configuredPrice")
                 or _s(v,"retailPrice") or _s(v,"grossListPrice") or {})
    rrp = (_num(price_obj) if not isinstance(price_obj, dict)
           else _num(price_obj.get("grossAmount") or price_obj.get("amount")
                     or price_obj.get("value")))

    monthly = _num(_s(v,"monthlyRate") or _s(v,"installment") or _s(v,"rate"))
    apr     = str(_s(v,"apr") or _s(v,"effectiveInterestRate") or "")

    avail_raw = str(_s(v,"availability") or _s(v,"stockStatus") or _s(v,"stock","stockType") or "")
    availability = ("available" if any(x in avail_raw.lower()
                    for x in ["now","immediately","stock","available"])
                    else ("soon" if avail_raw else ""))
    published_date = (_str(_s(v, "publishedDate"))
                      or _str(_s(v, "publicationDate"))
                      or _str(_s(v, "datePublished"))
                      or _str(_s(v, "onlineSince"))
                      or _str(_s(v, "stock", "publishedDate"))
                      or _str(_s(v, "stock", "publicationDate"))
                      or _str(_s(v, "stock", "createdDate"))
                      or _str(_s(hit, "publishedDate"))
                      or _str(_s(hit, "publicationDate"))
                      or _str(_s(hit, "datePublished"))
                      or "")

    dealer_obj  = _s(v,"dealer") or _s(hit,"dealer") or {}
    dealer_name = (_str(_s(v,"dealerName"))
                   or _str(_s(dealer_obj,"name"))
                   or _str(_s(dealer_obj,"nameLocalLanguage"))
                   or _str(_s(dealer_obj,"displayName"))
                   or "")
    city        = (_str(_s(v,"city"))
                   or _str(_s(dealer_obj,"city"))
                   or _str(_s(dealer_obj,"location","city"))
                   or "")
    if not city:
        address = _str(_s(dealer_obj, "addressLocalLanguage"))
        if address:
            city = address.split(",")[-1].strip().split(" ", 1)[-1]

    raw_id = (_s(v,"carId") or _s(v,"id") or _s(v,"vin")
              or _s(v,"vehicleId") or _s(v,"baumuster")
              or _s(v,"identification","code")
              or _s(v,"identification","variantId")
              or _s(hit,"id") or "")
    car_id = str(raw_id) if raw_id else ""

    car_url = (_str(_s(v,"detailUrl")) or _str(_s(v,"url"))
               or _str(_s(hit,"detailUrl")) or "")
    if car_url and not car_url.startswith("http"):
        car_url = "https://www.mercedes-benz.pt" + car_url
    if not car_url and car_id:
        car_url = (
            "https://www.mercedes-benz.pt/passengercars/buy/new-car/"
            "search-results.html/vehicleCategory-new-passenger-cars/details/"
            + car_id
        )

    return {
        "type":"New","model":model,"version":version,"body":body,
        "fuel":fuel,"gearbox":gearbox,"power":power,"drive":drive,
        "ext_color":ext_color,"int_color":int_color,
        "rrp_eur":rrp,"monthly_rate":monthly,"apr":apr,
        "availability":availability,"dealer":dealer_name,"city":city,
        "published_date":published_date,"car_id":car_id,"url":car_url,
    }


def enrich_prices(cars):
    session = requests.Session()
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
        "consumer-id": "VEHICLE_TILE",
        "Origin": "https://www.mercedes-benz.pt",
        "Referer": START_URL,
    })
    updated = 0
    for i, car in enumerate(cars, 1):
        car_id = car.get("car_id")
        if not car_id or car.get("rrp_eur"):
            continue
        try:
            resp = session.get(
                PRICE_API.format(car_id=car_id),
                params={"customerType": "B2C", "language": "pt"},
                timeout=20,
            )
            if resp.status_code != 200:
                continue
            data = resp.json()
            amount = _s(data, "purchasePrice", "amount")
            try:
                price = float(str(amount).replace(",", "."))
            except Exception:
                price = _num(amount)
            if price is not None:
                car["rrp_eur"] = price
                updated += 1
        except Exception:
            pass
        time.sleep(0.25)   # 250ms between requests — avoids rate limiting
        if i % 50 == 0:
            print(f"\r  Price enrichment: {i}/{len(cars)} | found: {updated}   ", end="", flush=True)
    print(f"\n  Price enrichment complete: {updated}/{len(cars)} prices filled")
    return cars


# ═══════════════════════════════════════════════════════
# COOKIE HANDLING
# ═══════════════════════════════════════════════════════

def accept_cookies(page):
    for label in ["Accept all","Aceitar tudo","Aceitar","Allow all","Confirmar"]:
        try:
            loc = page.get_by_text(label, exact=False).first
            if loc.count() > 0:
                loc.click(timeout=4000)
                page.wait_for_timeout(1500)
                print(f"   Cookies accepted: '{label}'")
                return
        except Exception:
            continue
    for sel in ["button#onetrust-accept-btn-handler","button.accept-all",
                "[data-testid='accept-all-cookies']","button[class*='accept']",
                "[id*='cookie'] button","button[class*='cookie']"]:
        try:
            el = page.locator(sel).first
            if el.count() > 0:
                el.click(timeout=4000)
                page.wait_for_timeout(1500)
                print(f"   Cookies accepted via: {sel}")
                return
        except Exception:
            continue
    print("   Warning: cookie banner not found")


async def accept_cookies_async(page):
    for label in ["Accept all", "Aceitar tudo", "Aceitar", "Allow all", "Confirmar"]:
        try:
            loc = page.get_by_text(label, exact=False).first
            if await loc.count() > 0:
                await loc.click(timeout=4000)
                await page.wait_for_timeout(1500)
                print(f"   Cookies accepted: '{label}'")
                return
        except Exception:
            continue
    for sel in ["button#onetrust-accept-btn-handler", "button.accept-all",
                "[data-testid='accept-all-cookies']", "button[class*='accept']",
                "[id*='cookie'] button", "button[class*='cookie']"]:
        try:
            el = page.locator(sel).first
            if await el.count() > 0:
                await el.click(timeout=4000)
                await page.wait_for_timeout(1500)
                print(f"   Cookies accepted via: {sel}")
                return
        except Exception:
            continue
    print("   Warning: cookie banner not found")


# ═══════════════════════════════════════════════════════
# MAIN SCRAPING LOGIC
# ═══════════════════════════════════════════════════════

def scrape_mercedes():
    try:
        asyncio.get_running_loop()
    except RuntimeError:
        pass
    else:
        raise RuntimeError("In Jupyter/VS Code notebooks use: raw_hits = await scrape_mercedes_async()")

    from playwright.sync_api import sync_playwright

    resp_queue  = queue.Queue()
    graphql_req = [None]
    total_count = [0]

    def on_request(request):
        try:
            if "graphql" not in request.url.lower():
                return
            body = request.post_data_json
            if not body:
                return
            # Only capture the vehicle search query (not misc other graphql calls)
            variables = body.get("variables") or {}
            if "profileId" not in variables and "vehicleCategory" not in variables:
                return
            graphql_req[0] = {
                "url":     request.url,
                "body":    body,
                "headers": {k: v for k, v in request.headers.items()
                            if k.lower() not in ("content-length","host",
                                                 "origin","referer","content-type")},
            }
            print(f"\n   GraphQL request captured: {request.url}")
        except Exception:
            pass

    def on_response(response):
        try:
            if "graphql" not in response.url.lower():
                return
            ct = response.headers.get("content-type","")
            if "json" not in ct:
                return
            resp_queue.put((response.url, response.text()))
        except Exception:
            pass

    with sync_playwright() as p:
        print("Opening browser...")
        headless = os.getenv("PLAYWRIGHT_HEADLESS", "0").lower() in ("1", "true", "yes")
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="pt-PT",
            viewport={"width": 1440, "height": 900},
        )
        page = ctx.new_page()
        page.on("request",  on_request)
        page.on("response", on_response)
        page.set_default_timeout(120000)

        print(f"Loading: {START_URL}")
        page.goto(START_URL, wait_until="domcontentloaded", timeout=120000)
        page.wait_for_timeout(3000)
        accept_cookies(page)
        page.wait_for_timeout(2000)

        # Wait for GraphQL request to be captured
        print(f"Waiting up to {CAPTURE_WAIT}s for GraphQL request...")
        for i in range(CAPTURE_WAIT):
            if graphql_req[0]:
                break
            page.wait_for_timeout(1000)
            if i % 5 == 4:
                page.evaluate("window.scrollBy(0, 400)")

        if not graphql_req[0]:
            print("  GraphQL request not captured. Check the browser manually.")
            browser.close()
            return []

        # Drain initial response
        first_hits = []
        first_total = 0
        page.wait_for_timeout(2000)
        while not resp_queue.empty():
            url, text = resp_queue.get_nowait()
            try:
                data = json.loads(text)
                hits, total = extract_hits(data)
                first_hits.extend(hits)
                if total: first_total = total
            except Exception:
                pass

        if first_total:
            total_count[0] = first_total
        print(f"\n   Page 0: {len(first_hits)} vehicles, total reported: {total_count[0]}")

        # Save raw first response for debugging
        debug_file = Path("mercedes_pt_raw_page0.json")
        try:
            # Re-fetch page 0 to save raw
            resp0 = page.request.post(
                graphql_req[0]["url"],
                data=json.dumps(graphql_req[0]["body"]),
                headers={**graphql_req[0]["headers"], "Content-Type": "application/json"},
                timeout=20000,
            )
            debug_file.write_text(resp0.text(), encoding="utf-8")
            print(f"   Raw page 0 saved -> {debug_file} (open to inspect JSON structure)")
        except Exception as e:
            print(f"   Could not save debug file: {e}")

        # === PAGINATION via direct GraphQL calls ===
        all_hits = list(first_hits)
        seen_ids = set()
        for h in all_hits:
            raw_id = _s(h,"carId") or _s(h,"id") or _s(h,"vin") or str(id(h))
            seen_ids.add(str(raw_id))

        gql      = graphql_req[0]
        total    = total_count[0] or 9999
        n_pages  = -(-total // PAGE_SIZE)   # ceil division
        print(f"\n   Paginating {n_pages} pages ({total} vehicles, {PAGE_SIZE}/page)...")

        consecutive_empty = 0
        for page_num in range(1, n_pages + 10):   # +10 safety margin
            if consecutive_empty >= 3:
                break
            # Patch body: set page variable
            body = json.loads(json.dumps(gql["body"]))
            variables = body.get("variables") or {}
            variables["page"] = page_num
            body["variables"] = variables

            try:
                resp = page.request.post(
                    gql["url"],
                    data=json.dumps(body),
                    headers={**gql["headers"], "Content-Type": "application/json"},
                    timeout=30000,
                )
                if resp.status >= 400:
                    print(f"\n   HTTP {resp.status} on page {page_num}, stopping.")
                    break
                data = json.loads(resp.text())
                hits, _ = extract_hits(data)
                new = 0
                for h in hits:
                    raw_id = _s(h,"carId") or _s(h,"id") or _s(h,"vin") or str(id(h))
                    vid = str(raw_id)
                    if vid not in seen_ids:
                        seen_ids.add(vid)
                        all_hits.append(h)
                        new += 1
                consecutive_empty = 0 if new else consecutive_empty + 1
                print(f"\r   Page {page_num:03d}/{n_pages} | New: {new:2d} | "
                      f"Total: {len(all_hits)}/{total}   ",
                      end="", flush=True)
                if len(all_hits) >= total:
                    break
                time.sleep(0.4)
            except Exception as e:
                print(f"\n   Error on page {page_num}: {e}")
                consecutive_empty += 1

        print(f"\n\n   Final hit count: {len(all_hits)}")
        browser.close()

    return all_hits


async def scrape_mercedes_async():
    from playwright.async_api import async_playwright

    captured_gql = {}
    all_hits = []
    seen_ids = set()

    async def on_request(request):
        try:
            if "graphql" not in request.url.lower():
                return
            body = request.post_data
            if not body:
                return
            parsed = json.loads(body)
            variables = parsed.get("variables") or {}
            if "profileId" not in variables and "vehicleCategory" not in variables:
                return
            captured_gql["url"] = request.url
            captured_gql["body"] = parsed
            captured_gql["headers"] = {
                k: v for k, v in (await request.all_headers()).items()
                if not k.startswith(":")
                and k.lower() not in ("content-length", "host", "origin", "referer", "content-type")
            }
            print(f"\n   GraphQL request captured: {request.url}")
        except Exception:
            pass

    async with async_playwright() as p:
        print("Opening browser...")
        headless = os.getenv("PLAYWRIGHT_HEADLESS", "0").lower() in ("1", "true", "yes")
        browser = await p.chromium.launch(headless=headless)
        try:
            ctx = await browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale="pt-PT",
                viewport={"width": 1440, "height": 900},
            )
            page = await ctx.new_page()
            page.on("request", lambda request: asyncio.create_task(on_request(request)))
            page.set_default_timeout(120000)

            print(f"Loading: {START_URL}")
            await page.goto(START_URL, wait_until="domcontentloaded", timeout=120000)
            await page.wait_for_timeout(3000)
            await accept_cookies_async(page)
            await page.wait_for_timeout(2000)

            print(f"Waiting up to {CAPTURE_WAIT}s for GraphQL request...")
            for i in range(CAPTURE_WAIT):
                if captured_gql:
                    break
                await page.wait_for_timeout(1000)
                if i % 5 == 4:
                    await page.evaluate("window.scrollBy(0, 400)")

            if not captured_gql:
                print("  GraphQL request not captured. Check the browser manually.")
                return []

            gql = captured_gql
            hdrs = {**gql["headers"], "Content-Type": "application/json"}

            resp0 = await page.request.post(
                gql["url"],
                data=json.dumps(gql["body"]),
                headers=hdrs,
                timeout=20000,
            )
            raw0 = await resp0.text()
            Path("mercedes_pt_raw_page0.json").write_text(raw0, encoding="utf-8")
            print("   Raw page 0 saved -> mercedes_pt_raw_page0.json")

            hits0, total_count = extract_hits(json.loads(raw0))
            print(f"\n   Page 0: {len(hits0)} vehicles, total reported: {total_count}")

            for h in hits0:
                raw_id = _s(h, "carId") or _s(h, "id") or _s(h, "vin") or str(id(h))
                vid = str(raw_id)
                if vid not in seen_ids:
                    seen_ids.add(vid)
                    all_hits.append(h)

            total = total_count or 9999
            n_pages = -(-total // PAGE_SIZE)
            print(f"\n   Paginating {n_pages} pages ({total} vehicles, {PAGE_SIZE}/page)...")

            consecutive_empty = 0
            for page_num in range(1, n_pages + 10):
                if consecutive_empty >= 3:
                    break
                body = json.loads(json.dumps(gql["body"]))
                variables = body.get("variables") or {}
                variables["page"] = page_num
                body["variables"] = variables

                try:
                    resp = await page.request.post(
                        gql["url"],
                        data=json.dumps(body),
                        headers=hdrs,
                        timeout=30000,
                    )
                    if resp.status >= 400:
                        print(f"\n   HTTP {resp.status} on page {page_num}, stopping.")
                        break
                    data = await resp.json()
                    hits, _ = extract_hits(data)
                    new = 0
                    for h in hits:
                        raw_id = _s(h, "carId") or _s(h, "id") or _s(h, "vin") or str(id(h))
                        vid = str(raw_id)
                        if vid not in seen_ids:
                            seen_ids.add(vid)
                            all_hits.append(h)
                            new += 1
                    consecutive_empty = 0 if new else consecutive_empty + 1
                    print(f"\r   Page {page_num:03d}/{n_pages} | New: {new:2d} | "
                          f"Total: {len(all_hits)}/{total}   ",
                          end="", flush=True)
                    if len(all_hits) >= total:
                        break
                    await asyncio.sleep(0.4)
                except Exception as e:
                    print(f"\n   Error on page {page_num}: {e}")
                    consecutive_empty += 1

            print(f"\n\n   Final hit count: {len(all_hits)}")
            return all_hits
        finally:
            await browser.close()


# ═══════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════

def save_excel(cars, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cars in stock"
    ws.freeze_panes = "A2"

    hdr_fill = PatternFill("solid", fgColor=MB_BLUE)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor=MB_LIGHT)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, car in enumerate(cars, 2):
        for col_i, field in enumerate(FIELD_MAP, 1):
            val = car.get(field, "")
            if val in ("", None) and field in ("rrp_eur","monthly_rate"):
                val = None
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if row_i % 2 == 0:
                cell.fill = alt_fill

    for col_i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = w
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Summary by Model")
    ws2.append(["Model", "Vehicles"])
    for model, count in Counter(c.get("model","?") for c in cars).most_common():
        ws2.append([model, count])
    for c in ws2[1]: c.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12

    wb.save(path)
    print(f"  Excel saved: {path}")


# ═══════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════

def main():
    start = datetime.now()
    print("=" * 65)
    print("  MERCEDES-BENZ PORTUGAL - STOCK SCRAPER")
    print(f"  {start.strftime('%Y-%m-%d %H:%M:%S')}")
    print("  Source: eu.api.oneweb.mercedes-benz.com GraphQL")
    print("=" * 65)

    raw_hits = asyncio.run(scrape_mercedes_async())

    if not raw_hits:
        print("\n  No vehicles found.")
        return

    print(f"\n[2/2] Parsing {len(raw_hits)} raw hits...")
    cars = [extract_vehicle(h) for h in raw_hits]
    cars = enrich_prices(cars)

    # Show what we extracted for the first 5 to sanity-check
    print("\n  First 5 extracted records:")
    for c in cars[:5]:
        print(f"    model={c.get('model')!r:30}  fuel={c.get('fuel')!r:15}  rrp={c.get('rrp_eur')}")

    cars = [c for c in cars if c.get("model")]

    elapsed = datetime.now() - start
    save_excel(cars, OUT_XLSX)

    print("\n" + "=" * 65)
    print("  FINAL SUMMARY")
    print(f"  Total vehicles : {len(cars)}")
    for m, n in Counter(c.get("model","?") for c in cars).most_common(10):
        print(f"    {m:<35} {n:>4} units")
    print(f"\n  Excel : {Path(OUT_XLSX).resolve()}")
    print(f"  Time  : {elapsed}")
    print("=" * 65)
    print("\n  NOTE: If model names look wrong, open mercedes_pt_raw_page0.json")
    print("  and share the structure so the field extraction can be adjusted.")


if __name__ == "__main__":
    main()
