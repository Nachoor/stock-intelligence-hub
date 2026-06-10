"""

scraper_mercedes_all_models.py
================================
Scraper de TODOS los modelos Mercedes-Benz en stock en España.
Fuente: stockmercedesbenz.es (concesionarios oficiales)

Instalar dependencias:
    pip install requests beautifulsoup4 openpyxl

Uso:
    python scraper_mercedes_all_models.py

Salida:
    STOCK_MERCEDES.xlsx                (en el directorio actual)
    dealers.json                        (lista de concesionarios encontrados)
"""

import requests
import json
import time
import urllib.parse
import re
import os
import sys
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ═══════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════

DELAY      = 0.5
timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
OUT_XLSX   = "STOCK_MERCEDES.xlsx"
OUT_JSON   = "dealers.json"

HUB_URL    = "https://stockmercedesbenz.es"
API_URL    = f"{HUB_URL}/concesionarios/"
SEED_URL   = f"{HUB_URL}/listado?carrocerias%5b%5d=6&modelos%5b%5d=24"

# (codigo_postal, lat, lng, ciudad) — una por provincia (50 provincias + Ceuta/Melilla)
# Bounds ±3.5 grados por llamada (~350 km); solapamiento intencional para no perder dealers
LOCATIONS = [
    # Andalucía
    ("04001", 36.8340,  -2.4637, "Almería"),
    ("11001", 36.5348,  -6.2998, "Cádiz"),
    ("14001", 37.8882,  -4.7794, "Córdoba"),
    ("18001", 37.1773,  -3.5986, "Granada"),
    ("21001", 37.2614,  -6.9447, "Huelva"),
    ("23001", 37.7796,  -3.7849, "Jaén"),
    ("29001", 36.7213,  -4.4214, "Málaga"),
    ("41001", 37.3886,  -5.9953, "Sevilla"),
    # Aragón
    ("22001", 42.1401,  -0.4089, "Huesca"),
    ("44001", 40.3456,  -1.1065, "Teruel"),
    ("50001", 41.6488,  -0.8891, "Zaragoza"),
    # Asturias
    ("33001", 43.3619,  -5.8494, "Oviedo"),
    # Baleares
    ("07001", 39.5696,   2.6502, "Palma"),
    # Canarias
    ("35001", 28.0997, -15.4134, "Las Palmas GC"),
    ("38001", 28.4636, -16.2518, "S.C. Tenerife"),
    # Cantabria
    ("39001", 43.4623,  -3.8099, "Santander"),
    # Castilla-La Mancha
    ("02001", 38.9943,  -1.8585, "Albacete"),
    ("13001", 38.9848,  -3.9274, "Ciudad Real"),
    ("16001", 40.0704,  -2.1374, "Cuenca"),
    ("19001", 40.6333,  -3.1661, "Guadalajara"),
    ("45001", 39.8628,  -4.0273, "Toledo"),
    # Castilla y León
    ("05001", 40.6564,  -4.6813, "Ávila"),
    ("09001", 42.3440,  -3.6970, "Burgos"),
    ("24001", 42.5987,  -5.5671, "León"),
    ("34001", 42.0096,  -4.5288, "Palencia"),
    ("37001", 40.9701,  -5.6635, "Salamanca"),
    ("40001", 40.9429,  -4.1088, "Segovia"),
    ("42001", 41.7640,  -2.4650, "Soria"),
    ("47001", 41.6523,  -4.7245, "Valladolid"),
    ("49001", 41.5036,  -5.7447, "Zamora"),
    # Cataluña
    ("08001", 41.3851,   2.1734, "Barcelona"),
    ("17001", 41.9792,   2.8214, "Girona"),
    ("25001", 41.6176,   0.6200, "Lleida"),
    ("43001", 41.1172,   1.2546, "Tarragona"),
    # Comunitat Valenciana
    ("03001", 38.3452,  -0.4815, "Alicante"),
    ("12001", 39.9864,  -0.0513, "Castellón"),
    ("46001", 39.4699,  -0.3763, "Valencia"),
    # Extremadura
    ("06001", 38.8794,  -6.9706, "Badajoz"),
    ("10001", 39.4753,  -6.3724, "Cáceres"),
    # Galicia
    ("15001", 43.3623,  -8.4115, "A Coruña"),
    ("27001", 43.0097,  -7.5567, "Lugo"),
    ("32001", 42.3364,  -7.8641, "Ourense"),
    ("36001", 42.4336,  -8.6453, "Pontevedra"),
    # La Rioja
    ("26001", 42.4650,  -2.4499, "Logroño"),
    # Madrid
    ("28001", 40.4260,  -3.6839, "Madrid"),
    # Murcia
    ("30001", 37.9922,  -1.1307, "Murcia"),
    # Navarra
    ("31001", 42.8125,  -1.6458, "Pamplona"),
    # País Vasco
    ("01001", 42.8467,  -2.6726, "Vitoria"),
    ("48001", 43.2630,  -2.9350, "Bilbao"),
    ("20001", 43.3183,  -1.9812, "San Sebastián"),
    # Ceuta y Melilla
    ("51001", 35.8894,  -5.3198, "Ceuta"),
    ("52001", 35.2923,  -2.9381, "Melilla"),
]


# ═══════════════════════════════════════════════════════
# SESIÓN HTTP
# ═══════════════════════════════════════════════════════

def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    })
    r = s.get(SEED_URL, timeout=20)
    r.raise_for_status()
    print(f"  Sesión establecida (status {r.status_code})")
    return s


def get_xsrf(session):
    return urllib.parse.unquote(session.cookies.get("XSRF-TOKEN", ""))


# ═══════════════════════════════════════════════════════
# DESCUBRIMIENTO DE CONCESIONARIOS
# ═══════════════════════════════════════════════════════

# Semillas: (carroceria, modelo, descripcion)
# Modelos populares que tienen la mayoría de concesionarios MB en stock
# carrocerias: 1=Berlina, 2=Estate, 3=Compacto, 4=Cabrio/Roadster, 6=SUV, 7=Coupé
# modelos: 1=ClaseA, 3=ClaseC, 8=GLC, 9=GLB, 11=GLE, 24=EQA, 27=EQE, 29=GLS
SEEDS = [
    ("6",  "8",  "GLC SUV"),
    ("6",  "11", "GLE SUV"),
    ("6",  "9",  "GLB SUV"),
    ("6",  "24", "EQA SUV"),
    ("1",  "3",  "Clase C Berlina"),
    ("1",  "1",  "Clase A Berlina"),
    ("2",  "3",  "Clase C Estate"),
    ("3",  "1",  "Clase A Compacto"),
    ("6",  "29", "GLS SUV"),
    ("7",  "3",  "Clase C Coupé"),
]


def get_all_dealers(session):
    xsrf = get_xsrf(session)
    seen = {}
    total_calls = 0

    for i, (cp, lat, lng, ciudad) in enumerate(LOCATIONS, 1):
        ne = json.dumps({"lat": round(lat + 3.5, 4), "lng": round(lng + 3.5, 4)})
        sw = json.dumps({"lat": round(lat - 3.5, 4), "lng": round(lng - 3.5, 4)})
        nuevos_loc = 0

        for carro, modelo, desc in SEEDS:
            params = [
                ("carrocerias[]",    carro),
                ("modelos[]",        modelo),
                ("origLat",          str(lat)),
                ("origLng",          str(lng)),
                ("origAddress",      cp),
                ("formattedAddress", f"{cp} {ciudad}, España"),
                ("boundsNorthEast",  ne),
                ("boundsSouthWest",  sw),
            ]
            try:
                r = session.get(API_URL, params=params, headers={
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "X-Requested-With": "XMLHttpRequest",
                    "X-XSRF-TOKEN": xsrf,
                    "Referer": SEED_URL,
                }, timeout=15)
                total_calls += 1

                if r.status_code == 200:
                    data = r.json()
                    if isinstance(data, list):
                        for d in data:
                            if d["id_grupo_concesionario"] not in seen:
                                seen[d["id_grupo_concesionario"]] = d
                                nuevos_loc += 1
            except Exception:
                pass
            time.sleep(DELAY)

        if nuevos_loc > 0:
            print(f"  [{i:02d}/{len(LOCATIONS)}] {ciudad} ({cp}) → +{nuevos_loc} nuevos (total {len(seen)})")

    print(f"\n  Llamadas API realizadas: {total_calls}")
    return list(seen.values())


# ═══════════════════════════════════════════════════════
# PARSING DE TARJETAS DE COCHES
# ═══════════════════════════════════════════════════════

def normalize_date_text(value):
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return text[:10] if re.match(r"\d{4}-\d{2}-\d{2}", text) else ""


def extract_publication_date(text):
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if not text:
        return ""
    date_re = r"(\d{4}-\d{2}-\d{2}|\d{1,2}[./-]\d{1,2}[./-]\d{2,4})"
    for match in re.finditer(date_re, text):
        start = max(0, match.start() - 80)
        context = text[start:match.end() + 40].lower()
        if any(token in context for token in ("publicad", "alta", "online", "desde", "fecha")):
            return normalize_date_text(match.group(1))
    return ""

def parse_page(html, dealer_name, dealer_city, dealer_state, base_domain):
    """
    Extrae todos los coches de una página HTML del showroom.
    - Página completa /b/  → sin separador, se parsea directamente
    - Fragmento AJAX /bl   → coches están DESPUÉS de <!--@-->
    """
    if "<!--@-->" in html:
        html = html.split("<!--@-->")[1]

    soup = BeautifulSoup(html, "html.parser")
    cars = []

    for card in soup.select("div.col-lg-4.col-md-6.col-sm-12.p-3"):
        try:
            link = card.find("a", class_="microche")
            if not link:
                continue

            car_path = link.get("href", "")
            if car_path.startswith("/"):
                car_url = f"https://{base_domain}{car_path}"
            elif car_path.startswith("http"):
                car_url = car_path
            else:
                car_url = f"https://{base_domain}/{car_path}"

            img     = card.find("img", {"data-version": True})
            color   = img.get("data-color", "").strip() if img else ""
            version = img.get("data-version", "").strip() if img else ""

            status_el = card.find("div", class_="status")
            estado    = status_el.get_text(strip=True) if status_el else ""

            tipo = "Ocasión" if card.find(class_=lambda c: c and "ocas" in c.lower()) else "Nuevo"

            model_el = card.find("p", class_=lambda c: c and "fuente2" in c)
            modelo   = model_el.get_text(strip=True) if model_el else version

            pvp_el   = card.find("p", class_=lambda c: c and "fuente1" in c)
            pvp_text = pvp_el.get_text(strip=True) if pvp_el else ""

            pvp_m  = re.search(r"PVP\s*:?\s*([\d.,]+)\s*€", pvp_text, re.IGNORECASE)
            pvp    = pvp_m.group(1).replace(".", "").replace(",", ".") if pvp_m else ""

            cuota_m = re.search(r"(?:\||Renting\s*:?)\s*([\d.,]+)\s*€/mes", pvp_text, re.IGNORECASE)
            cuota   = cuota_m.group(1).replace(".", "").replace(",", ".") if cuota_m else ""

            motor_el = card.find("p", class_=lambda c: c and "bt-gris" in c)
            motor    = motor_el.get_text(strip=True).rstrip(",") if motor_el else ""
            fecha_publicacion = extract_publication_date(card.get_text(" ", strip=True))

            tae = ""
            for p in card.find_all("p", class_=lambda c: c and "fuente1" in c):
                tae_m = re.search(r"TAE:\s*([\d.,]+)%", p.get_text(strip=True))
                if tae_m:
                    tae = tae_m.group(1)
                    break

            cars.append({
                "concesionario": dealer_name,
                "ciudad":        dealer_city,
                "provincia":     dealer_state,
                "tipo":          tipo,
                "modelo":        modelo,
                "version":       version,
                "color":         color,
                "pvp_eur":       pvp,
                "cuota_mes_eur": cuota,
                "tae_pct":       tae,
                "motor":         motor,
                "estado":        estado,
                "fecha_publicacion": fecha_publicacion,
                "url_coche":     car_url,
            })

        except Exception:
            continue

    return cars


# ═══════════════════════════════════════════════════════
# SCRAPING DE UN CONCESIONARIO (TODAS LAS PÁGINAS)
# ═══════════════════════════════════════════════════════

def scrape_dealer(session, dealer):
    """
    Scrapea TODOS los coches de un concesionario sin filtro de modelo.
    Pág 1 → /b/ (HTML completo, server-rendered)
    Pág 2+ → /bl?page=N (fragmento AJAX)
    """
    url_landing = dealer.get("url_landing", "")
    if not url_landing:
        return []

    full_url = url_landing if url_landing.startswith("http") else f"https://{url_landing}"
    domain   = urllib.parse.urlparse(full_url).netloc

    name  = dealer["name"]
    city  = dealer["city"]
    state = dealer["state"]

    all_cars = []
    total    = 9999
    page     = 1

    while True:
        if page == 1:
            page_url = f"https://{domain}/b/"
            hdrs = {
                "Accept": "text/html,application/xhtml+xml,*/*",
                "Referer": f"https://{domain}/b/",
            }
        else:
            page_url = f"https://{domain}/bl?page={page}"
            hdrs = {
                "Accept": "text/html, */*; q=0.01",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": f"https://{domain}/b/",
            }

        try:
            r = session.get(page_url, headers=hdrs, timeout=20)

            if r.status_code != 200:
                break

            html = r.text

            # Actualizar total si está disponible
            count_m = re.search(r'data-count="(\d+)"', html)
            if count_m:
                total = int(count_m.group(1))

            cars = parse_page(html, name, city, state, domain)
            if not cars:
                break

            all_cars.extend(cars)

            if len(all_cars) >= total:
                break

            # Comprobar si hay más páginas
            frag = html.split("<!--@-->")[1] if "<!--@-->" in html else html
            soup = BeautifulSoup(frag, "html.parser")
            pagination = soup.find("ul", class_="pagination")
            if pagination:
                page_nums = []
                for a in pagination.find_all("a", class_="page-link"):
                    try:
                        page_nums.append(int(a.get_text(strip=True)))
                    except ValueError:
                        pass
                if page >= max(page_nums, default=page):
                    break
            else:
                break

            page += 1
            time.sleep(DELAY)

        except Exception as e:
            print(f"      ⚠ Error página {page}: {e}")
            break

    return all_cars


def enrich_missing_prices_from_detail(session, cars):
    """Fill missing PVP/monthly values from detail pages when cards only show renting."""
    targets = [c for c in cars if c.get("url_coche") and (not c.get("pvp_eur") or not c.get("cuota_mes_eur"))]
    if not targets:
        return cars

    updated_pvp = 0
    updated_monthly = 0
    for i, car in enumerate(targets, 1):
        try:
            r = session.get(car["url_coche"], timeout=20)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            text = soup.get_text(" ", strip=True)
            if not car.get("fecha_publicacion"):
                car["fecha_publicacion"] = extract_publication_date(text)

            if not car.get("pvp_eur"):
                pvp_m = re.search(r"PVP\s*:?\s*([\d.,]+)\s*€", text, re.IGNORECASE)
                if pvp_m:
                    car["pvp_eur"] = pvp_m.group(1).replace(".", "").replace(",", ".")
                    updated_pvp += 1

            if not car.get("cuota_mes_eur"):
                cuota_m = re.search(r"(?:Renting\s*:?\s*|\|\s*)([\d.,]+)\s*€/mes", text, re.IGNORECASE)
                if cuota_m:
                    car["cuota_mes_eur"] = cuota_m.group(1).replace(".", "").replace(",", ".")
                    updated_monthly += 1
        except Exception:
            pass

        if i % 25 == 0:
            print(f"\r  Enriqueciendo fichas Mercedes: {i}/{len(targets)} "
                  f"| PVP +{updated_pvp} | Cuotas +{updated_monthly}", end="", flush=True)
        time.sleep(DELAY)

    print(f"\n  Enriquecimiento fichas: PVP +{updated_pvp}, cuotas +{updated_monthly}")
    return cars


# ═══════════════════════════════════════════════════════
# EXPORTAR A EXCEL
# ═══════════════════════════════════════════════════════

HEADERS = [
    "Concesionario", "Ciudad", "Provincia", "Tipo",
    "Modelo", "Versión", "Color",
    "PVP (€)", "Cuota/mes (€)", "TAE (%)",
    "Motor/Combustible", "Estado", "Fecha Publicacion", "URL Coche",
]

FIELD_MAP = [
    "concesionario", "ciudad", "provincia", "tipo",
    "modelo", "version", "color",
    "pvp_eur", "cuota_mes_eur", "tae_pct",
    "motor", "estado", "fecha_publicacion", "url_coche",
]

COL_WIDTHS = [30, 16, 16, 10, 40, 45, 14, 12, 14, 8, 30, 20, 18, 65]


def save_excel(cars, dealers, path):
    from collections import Counter
    wb = openpyxl.Workbook()

    # ── Hoja 1: todos los coches ──
    ws = wb.active
    ws.title = "Coches en stock"
    ws.freeze_panes = "A2"

    hdr_fill = PatternFill("solid", fgColor="00ADEF")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="EAF6FF")

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, car in enumerate(cars, 2):
        for col_i, field in enumerate(FIELD_MAP, 1):
            raw = car.get(field, "")
            if field in ("pvp_eur", "cuota_mes_eur", "tae_pct"):
                try:
                    val = float(raw) if raw else None
                except ValueError:
                    val = raw
            else:
                val = raw
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if row_i % 2 == 0:
                cell.fill = alt_fill

    for col_i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = w

    ws.auto_filter.ref = ws.dimensions

    # ── Hoja 2: resumen por concesionario ──
    ws2 = wb.create_sheet("Resumen concesionarios")
    ws2.append(["Concesionario", "Ciudad", "Provincia", "Coches scrapeados", "URL Showroom"])

    counts = Counter(c["concesionario"] for c in cars)
    for d in sorted(dealers, key=lambda x: (x["state"], x["city"])):
        domain = urllib.parse.urlparse("https://" + d["url_landing"]).netloc
        ws2.append([
            d["name"], d["city"], d["state"],
            counts.get(d["name"], 0),
            f"https://{domain}/b/",
        ])

    for col in ws2.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 20
    ws2.column_dimensions["E"].width = 55

    wb.save(path)
    print(f"  ✓ Excel guardado: {path}")


# ═══════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════

def main():
    start = datetime.now()
    print("=" * 65)
    print("  SCRAPER MERCEDES-BENZ — TODOS LOS MODELOS EN STOCK")
    print(f"  {start.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)

    print("\n[1/3] Estableciendo sesión con stockmercedesbenz.es ...")
    session = make_session()

    print("\n[2/3] Descubriendo concesionarios con stock ...")
    dealers = get_all_dealers(session)
    print(f"\n  → {len(dealers)} concesionarios únicos encontrados")

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(dealers, f, ensure_ascii=False, indent=2)
    print(f"  → Lista guardada en {OUT_JSON}")

    print("\n[3/3] Scrapeando stock completo (todos los modelos) ...")
    all_cars = []

    for i, dealer in enumerate(dealers, 1):
        print(f"  [{i:02d}/{len(dealers)}] {dealer['name']} ({dealer['city']}) ...",
              end=" ", flush=True)
        cars = scrape_dealer(session, dealer)
        all_cars.extend(cars)
        print(f"{len(cars)} coches")

    print("\nEnriqueciendo precios/cuotas desde fichas ...")
    all_cars = enrich_missing_prices_from_detail(session, all_cars)

    elapsed = datetime.now() - start
    print(f"\n  ✓ Total coches scrapeados: {len(all_cars)}")
    print(f"  ✓ Tiempo total: {elapsed}")

    print("\nGuardando Excel ...")
    save_excel(all_cars, dealers, OUT_XLSX)

    print("\n" + "=" * 65)
    print("  RESUMEN FINAL")
    print(f"  Concesionarios : {len(dealers)}")
    print(f"  Coches en stock: {len(all_cars)}")
    print(f"  Archivo Excel  : {os.path.abspath(OUT_XLSX)}")
    print(f"  Tiempo         : {elapsed}")
    print("=" * 65)


if __name__ == "__main__":
    main()
