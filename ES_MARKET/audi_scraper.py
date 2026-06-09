"""
scraper_audi_stock.py
======================
Scraper de TODOS los Audi en stock en España.
Fuente: API oficial scs.audi.de (Stock Car Search)

Instalar dependencias:
    pip install requests openpyxl

Uso:
    python scraper_audi_stock.py

Salida:
    STOCK_AUDI.xlsx                 (en el directorio actual)
    audi_raw_YYYYMMDD_HHMM.json     (datos crudos, opcional)

Sin selección de CP ni concesionario — la API devuelve directamente
todos los vehículos nuevos de España de una vez.
"""

import requests
import json
import time
import os
from datetime import datetime
from urllib.parse import parse_qs, urlparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

# ═══════════════════════════════════════════════════════
# CONFIGURACIÓN
# ═══════════════════════════════════════════════════════

SCS_BASE    = "https://scs.audi.de/api"
SCS_VERSION = "v2"
SCS_MARKET  = "es/es"
API_TOKEN   = "FJ54W6H"          # Token público del buscador oficial audi.es
PAGE_SIZE   = 200                 # Máximo soportado por la API
DELAY       = 0.4                 # segundos entre páginas

timestamp  = datetime.now().strftime("%Y%m%d_%H%M")
OUT_XLSX   = "STOCK_AUDI.xlsx"
OUT_JSON   = f"audi_raw_{timestamp}.json"   # pon None para no guardarlo
SAVE_JSON  = True


def make_session():
    s = requests.Session()
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Referer": "https://www.audi.es/",
        "Token": API_TOKEN,
    })
    return s


# ═══════════════════════════════════════════════════════
# EXTRACCIÓN DE CAMPOS
# ═══════════════════════════════════════════════════════

def safe_get(obj, *keys, default=""):
    """Navega por claves anidadas de forma segura."""
    for k in keys:
        if obj is None:
            return default
        if isinstance(obj, dict):
            obj = obj.get(k)
        elif isinstance(obj, list):
            try:
                obj = obj[k]
            except (IndexError, TypeError):
                return default
        else:
            return default
    return obj if obj is not None else default


def to_str(val):
    """Convierte cualquier valor a string legible (extrae 'description' si es dict)."""
    if val is None or val == "":
        return ""
    if isinstance(val, dict):
        return val.get("description") or val.get("code") or str(val)
    return str(val)


def audi_detail_url(web_link):
    """Convert Audi entry links to the final stock detail URL."""
    if not web_link:
        return ""
    if "entry.audi.com" not in web_link:
        return web_link
    vehicle_id = parse_qs(urlparse(web_link).query).get("id", [""])[0]
    if not vehicle_id:
        return web_link
    return "https://www.audi.es/es/buscador-de-stock-nuevo/details/?vehicleid=" + vehicle_id


def parse_vehicle(v):
    """Extrae los campos relevantes de un vehicleBasic."""

    # Modelo
    modelo      = to_str(safe_get(v, "model", "description"))
    version     = to_str(safe_get(v, "modelVersion"))
    carline     = to_str(safe_get(v, "symbolicCarline", "description") or safe_get(v, "symbolicCarline"))
    year        = safe_get(v, "modelYear")
    trimline    = to_str(safe_get(v, "trimline"))
    body        = to_str(safe_get(v, "bodyType", "description"))

    # Mecánica
    fuel        = to_str(safe_get(v, "fuel", "description") or safe_get(v, "fuel"))
    gear        = to_str(safe_get(v, "gearType"))
    power       = to_str(safe_get(v, "powerDisplay"))
    drive       = to_str(safe_get(v, "driveType"))

    # Colores
    ext_color   = to_str(safe_get(v, "extColor", "description") or safe_get(v, "extColor"))
    ext_code    = to_str(safe_get(v, "extColor", "code"))
    int_color   = to_str(safe_get(v, "upholsteryColor"))
    int_type    = to_str(safe_get(v, "upholsteryType"))

    # Precio
    pvp         = safe_get(v, "priceDetails", "totalPrices", "customer", "grossAmount")
    pvp_label   = safe_get(v, "priceDetails", "totalPrices", "customer", "metaInfo", "label")
    cuota_rate  = safe_get(v, "typedPrices", "rate")
    # financing object
    fin         = v.get("financing") or {}
    tae         = safe_get(fin, "tae")
    tin         = safe_get(fin, "tin")
    entrada     = safe_get(fin, "deposit")
    meses       = safe_get(fin, "term")

    # Disponibilidad
    avail_code  = safe_get(v, "availableFromCode")   # "now" | "soon" | ...
    avail_text  = safe_get(v, "availableFrom")
    avail_date  = safe_get(v, "availableFromDate")
    if avail_date:
        try:
            avail_date = datetime.utcfromtimestamp(avail_date / 1000).strftime("%Y-%m-%d")
        except:
            avail_date = ""

    # Dealer
    dealer_city = to_str(safe_get(v, "dealer", "city"))
    dealer_name = ""
    dealer_id   = ""
    dlr_list    = safe_get(v, "dealer", "dealerContextLinkData") or []
    for dlr in dlr_list:
        if dlr.get("context") == "A" and dlr.get("groupName"):
            dealer_name = to_str(dlr["groupName"])
            dealer_id   = to_str(dlr.get("dealerId", ""))
            break
    if not dealer_name and dlr_list:
        dealer_name = to_str(safe_get(dlr_list, 0, "groupName"))
        dealer_id   = to_str(safe_get(dlr_list, 0, "dealerId"))

    dealer_email = to_str(safe_get(v, "dealer", "email"))

    # IDs y links
    car_id      = to_str(v.get("carId", ""))
    web_link    = to_str(v.get("weblink", ""))
    if web_link and not web_link.startswith("http"):
        web_link = "https://www.audi.es" + web_link
    web_link = audi_detail_url(web_link)
    tipo        = "Ocasión" if v.get("used") else "Nuevo"
    biz_model   = safe_get(v, "businessModel", "description")
    campaign    = "Sí" if v.get("hasCampaigns") else ""
    buyable     = "Sí" if v.get("buyableOnline") else ""

    return {
        "tipo":           tipo,
        "modelo":         modelo,
        "version":        version,
        "carline":        carline,
        "año":            year,
        "trimline":       trimline,
        "carroceria":     body,
        "combustible":    fuel,
        "cambio":         gear,
        "potencia":       power,
        "traccion":       drive,
        "color_ext":      ext_color,
        "codigo_color":   ext_code,
        "color_int":      int_color,
        "tapizado":       int_type,
        "pvp_eur":        float(pvp) if pvp else None,
        "pvp_label":      pvp_label,
        "cuota_mes_eur":  float(cuota_rate) if cuota_rate else None,
        "tae":            float(tae) if tae else None,
        "tin":            float(tin) if tin else None,
        "entrada_eur":    float(entrada) if entrada else None,
        "plazo_meses":    int(meses) if meses else None,
        "disponibilidad": avail_code,
        "entrega_texto":  str(avail_text)[:80] if avail_text else "",
        "fecha_disponible": avail_date,
        "concesionario":  dealer_name,
        "ciudad":         dealer_city,
        "dealer_id":      dealer_id,
        "email_dealer":   dealer_email,
        "car_id":         car_id,
        "con_campaña":    campaign,
        "compra_online":  buyable,
        "modelo_negocio": biz_model,
        "url_coche":      web_link,
    }


# ═══════════════════════════════════════════════════════
# SCRAPING
# ═══════════════════════════════════════════════════════

def fetch_all_vehicles(session):
    """Pagina por la API SCS y devuelve todos los vehicleBasic."""
    url = f"{SCS_BASE}/{SCS_VERSION}/search/filter/{SCS_MARKET}"
    params_base = {"sort": "prices.retail:asc", "size": PAGE_SIZE}

    # Primer request para saber el total
    r = session.get(url, params={**params_base, "from": 0}, timeout=20)
    r.raise_for_status()
    first = r.json()
    total = first.get("totalCount", 0)
    print(f"  Total vehículos en API: {total}")

    all_raw = first.get("vehicleBasic", [])
    print(f"  Página 1: {len(all_raw)} coches (from=0)")

    pages = -(-total // PAGE_SIZE)  # ceil division

    for page in range(1, pages):
        from_idx = page * PAGE_SIZE
        time.sleep(DELAY)
        try:
            r = session.get(url, params={**params_base, "from": from_idx}, timeout=20)
            r.raise_for_status()
            batch = r.json().get("vehicleBasic", [])
            all_raw.extend(batch)
            print(f"  Página {page+1}/{pages}: {len(batch)} coches (from={from_idx}, total acum={len(all_raw)})")
            if not batch:
                break
        except Exception as e:
            print(f"  ⚠ Error en página {page+1} (from={from_idx}): {e}")

    return all_raw, total


# ═══════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════

HEADERS = [
    "Tipo", "Modelo", "Versión", "Carline", "Año", "Trim",
    "Carrocería", "Combustible", "Cambio", "Potencia", "Tracción",
    "Color Ext.", "Código Color", "Color Int.", "Tapizado",
    "PVP (€)", "Etiqueta Precio", "Cuota/mes (€)", "TAE (%)", "TIN (%)",
    "Entrada (€)", "Plazo (meses)",
    "Disponibilidad", "Texto Entrega", "Fecha Disponible",
    "Concesionario", "Ciudad", "ID Dealer", "Email Dealer",
    "Car ID", "Campaña", "Online", "Modelo Negocio", "URL Coche",
]

FIELD_MAP = [
    "tipo", "modelo", "version", "carline", "año", "trimline",
    "carroceria", "combustible", "cambio", "potencia", "traccion",
    "color_ext", "codigo_color", "color_int", "tapizado",
    "pvp_eur", "pvp_label", "cuota_mes_eur", "tae", "tin",
    "entrada_eur", "plazo_meses",
    "disponibilidad", "entrega_texto", "fecha_disponible",
    "concesionario", "ciudad", "dealer_id", "email_dealer",
    "car_id", "con_campaña", "compra_online", "modelo_negocio", "url_coche",
]

WIDTHS = [
    8, 35, 40, 14, 6, 16,
    14, 14, 12, 10, 10,
    18, 14, 16, 12,
    12, 20, 14, 8, 8,
    12, 12,
    14, 50, 14,
    35, 20, 10, 28,
    20, 8, 8, 16, 65,
]

AUDI_GRAY  = "4E4E4E"
AUDI_LIGHT = "F0F0F0"


def save_excel(cars, path):
    wb = openpyxl.Workbook()

    # ── Hoja 1: todos los coches ──────────────────────────
    ws = wb.active
    ws.title = "Coches en stock"
    ws.freeze_panes = "A2"

    hdr_fill = PatternFill("solid", fgColor=AUDI_GRAY)
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor=AUDI_LIGHT)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row_i, car in enumerate(cars, 2):
        for col_i, field in enumerate(FIELD_MAP, 1):
            val = car.get(field, "")
            # Empty string → None for numeric columns
            if val == "" and field in ("pvp_eur", "cuota_mes_eur", "tae", "tin", "entrada_eur", "plazo_meses", "año"):
                val = None
            # Safety net: convert any remaining dict/list to string
            if isinstance(val, (dict, list)):
                val = to_str(val) if isinstance(val, dict) else str(val)
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if row_i % 2 == 0:
                cell.fill = alt_fill

    for col_i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = w

    ws.auto_filter.ref = ws.dimensions

    # ── Hoja 2: resumen por modelo ─────────────────────────
    ws2 = wb.create_sheet("Resumen por modelo")
    ws2.append(["Carline / Modelo", "Unidades en stock"])

    cnt_model = Counter(
        str(c.get("carline") or c.get("modelo") or "?").split(" ")[0]
        for c in cars
    )
    for model, count in sorted(cnt_model.items(), key=lambda x: -x[1]):
        ws2.append([model, count])

    for col in ws2.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 18

    # ── Hoja 3: resumen por concesionario ──────────────────
    ws3 = wb.create_sheet("Resumen por concesionario")
    ws3.append(["Concesionario", "Ciudad", "Coches en stock"])

    cnt_dealer = Counter(
        (c.get("concesionario", "?"), c.get("ciudad", ""))
        for c in cars
    )
    for (name, city), count in sorted(cnt_dealer.items(), key=lambda x: -x[1]):
        ws3.append([name, city, count])

    for col in ws3.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 18

    wb.save(path)
    print(f"  ✓ Excel guardado: {path}")


# ═══════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════

def main():
    start = datetime.now()
    print("=" * 65)
    print("  SCRAPER AUDI ESPAÑA — STOCK DE VEHÍCULOS NUEVOS")
    print(f"  {start.strftime('%Y-%m-%d %H:%M:%S')}")
    print("  Fuente: scs.audi.de (Stock Car Search API oficial)")
    print("=" * 65)

    print("\n[1/3] Conectando con la API SCS de Audi...")
    session = make_session()
    # Quick test
    r = session.get(
        f"{SCS_BASE}/{SCS_VERSION}/search/filter/{SCS_MARKET}",
        params={"from": 0, "size": 1, "sort": "prices.retail:asc"},
        timeout=10
    )
    r.raise_for_status()
    print(f"  ✓ API accesible (status {r.status_code})")

    print("\n[2/3] Descargando todos los vehículos...")
    raw_vehicles, total = fetch_all_vehicles(session)
    elapsed_fetch = datetime.now() - start
    print(f"\n  ✓ {len(raw_vehicles)}/{total} vehículos descargados ({elapsed_fetch})")

    print("\n[3/3] Parseando y exportando...")

    # Save raw JSON (optional)
    if SAVE_JSON:
        with open(OUT_JSON, "w", encoding="utf-8") as f:
            json.dump(raw_vehicles, f, ensure_ascii=False)
        print(f"  ✓ JSON crudo guardado: {OUT_JSON}")

    # Parse
    cars = [parse_vehicle(v) for v in raw_vehicles]

    # Summary stats
    modelos = Counter(str(c.get("carline") or "?").split(" ")[0] for c in cars)
    disponibles = Counter(c.get("disponibilidad", "?") for c in cars)

    # Save Excel
    save_excel(cars, OUT_XLSX)

    elapsed = datetime.now() - start
    print("\n" + "=" * 65)
    print("  RESUMEN FINAL")
    print(f"  Total coches: {len(cars)}")
    print(f"  Disponibles ya:  {disponibles.get('now', 0)}")
    print(f"  Próximamente:    {disponibles.get('soon', 0)}")
    print(f"  Top modelos:")
    for m, n in modelos.most_common(8):
        print(f"    {m:<20} {n:>5} uds")
    print(f"\n  Archivo Excel: {os.path.abspath(OUT_XLSX)}")
    print(f"  Tiempo total:  {elapsed}")
    print("=" * 65)


if __name__ == "__main__":
    main()
